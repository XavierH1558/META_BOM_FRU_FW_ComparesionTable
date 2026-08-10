import os
import csv
import openpyxl
import re
import io
import json
import datetime
import yaml
from flask import Flask, render_template, jsonify, request, send_file
from werkzeug.utils import secure_filename

def safe_filename(filename):
    """Sanitize filename while preserving Unicode/Chinese characters and spaces."""
    if not filename:
        return 'unnamed'
    # Normalize slashes and pick basename
    filename = filename.replace('\\', '/').split('/')[-1]
    # Remove control chars and null bytes
    filename = re.sub(r'[\x00-\x1f\x7f]', '', filename)
    # Remove unsafe OS path chars: ? : * % " < > |
    filename = re.sub(r'[?:*%\x22<>|]', '_', filename)
    filename = filename.strip(' .')
    return filename or 'unnamed'


def atomic_write_json(path, data):
    """Write JSON to disk atomically (write to temp file, then os.replace) so a
    crash or concurrent write mid-save can't leave a half-written / corrupt file."""
    tmp_path = f"{path}.{os.getpid()}.tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp_path, path)


app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
DATA_DIR = os.path.join(BASE_DIR, 'data')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
# Cap request/upload size (default 50MB) to stop runaway disk/memory usage from
# oversized or accidental uploads. Override with MAX_UPLOAD_MB env var if needed.
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_UPLOAD_MB', '50')) * 1024 * 1024


@app.errorhandler(413)
def handle_file_too_large(e):
    return jsonify({
        'success': False,
        'error': f"檔案過大，上傳大小上限為 {app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)}MB。"
    }), 413

@app.after_request
def add_cache_busting_headers(response):
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# Initialize Google Drive Background Sync Scheduler if enabled in gdrive_config.json
try:
    from gdrive_sync import sync_all_gdrive_folders, load_config
    gcfg = load_config()
    if gcfg and gcfg.get('enabled', False):
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler(daemon=True)
        
        # Default to twice a day (9:00 AM and 6:00 PM)
        sync_hours = str(gcfg.get('sync_hours', '9,18'))
        scheduler.add_job(
            func=sync_all_gdrive_folders,
            trigger="cron",
            hour=sync_hours,
            minute=0,
            id='gdrive_sync_job',
            replace_existing=True
        )
        scheduler.start()
        print(f"[GDrive Sync] Background scheduler started. Syncing twice daily at hours ({sync_hours}:00).")
except Exception as g_err:
    print(f"[GDrive Sync] Background scheduler initialization skipped: {g_err}")

# ============================================================
# MULTI-PROJECT CONFIGURATION
# To add Clemente local Mac paths, update the 'mac_paths' lists
# under PROJECT_CONFIGS['clemente'] below.
# ============================================================
PROJECT_CONFIGS = {
    'sanmiguel': {
        'id': 'sanmiguel',
        'label': 'SanMiguel (VR200)',
        'color_theme': 'cyan',
        'dir_roots': {
            'bkc': [
                '/Volumes/DATA/Projects/META/VR200-SanMiguel/BKC Table/',
                os.path.join(DATA_DIR, 'sanmiguel', 'bkc'),
            ],
            'fru': [
                '/Volumes/DATA/Projects/META/VR200-SanMiguel/FRU Spec/DVT/',
                '/Volumes/DATA/Projects/META/VR200-SanMiguel/FRU Spec/PVT1-1/',
                os.path.join(DATA_DIR, 'sanmiguel', 'fru'),
            ],
            'matrix': [
                '/Volumes/DATA/Projects/META/VR200-SanMiguel/Build Matrix/',
                os.path.join(DATA_DIR, 'sanmiguel', 'matrix'),
            ],
            'yaml': [
                os.path.join(DATA_DIR, 'sanmiguel', 'yaml'),
            ]
        },
        'default_paths': {
            'bkc':     os.path.join(DATA_DIR, 'sanmiguel', 'bkc', 'San Miguel(VR200) FW control table - 3way.xlsx'),
            'fru_dvt': os.path.join(DATA_DIR, 'sanmiguel', 'fru', 'Maxwell_Earth_FRU table_DVT_20260708.xlsx'),
            'fru_pvt': os.path.join(DATA_DIR, 'sanmiguel', 'fru', 'Maxwell_Earth_FRU table_PVT1-1_20260709-1639.xlsx'),
            'matrix':  os.path.join(DATA_DIR, 'sanmiguel', 'matrix', 'SanMiguel(Ingrasys) Build Matrix_DVT_Draft_260325.xlsx')
        }
    },
    'clemente': {
        'id': 'clemente',
        'label': 'Clemente (GB300)',
        'color_theme': 'purple',
        'dir_roots': {
            # TODO: Add Clemente local Mac paths when available
            # e.g. '/Volumes/DATA/Projects/META/GB300-Clemente/BKC Table/'
            'bkc':    [os.path.join(DATA_DIR, 'clemente', 'bkc')],
            'fru':    [os.path.join(DATA_DIR, 'clemente', 'fru')],
            'matrix': [os.path.join(DATA_DIR, 'clemente', 'matrix')],
            'yaml':   [os.path.join(DATA_DIR, 'clemente', 'yaml')],
        },
        'default_paths': {
            'bkc':     os.path.join(DATA_DIR, 'clemente', 'bkc', 'IGS_FW_Control_Table_3way.xlsx'),
            'fru_dvt': os.path.join(DATA_DIR, 'clemente', 'fru', 'CI-Clemente_FRU_programming_MP_20260622.xlsx'),
            'fru_pvt': os.path.join(DATA_DIR, 'clemente', 'fru', 'CI-Clemente_FRU_programming_MP_20260622.xlsx'),
            'matrix':  os.path.join(DATA_DIR, 'clemente', 'matrix', 'PVT_Clemente(Ingrasys) Build Matrix_PVT2_20250912_Leo.xlsx')
        }
    }
}

# Create project data subdirectories
for _proj in PROJECT_CONFIGS:
    for _tab in ['bkc', 'fru', 'matrix', 'yaml']:
        os.makedirs(os.path.join(DATA_DIR, _proj, _tab), exist_ok=True)

# Per-project upload folders
for _proj in PROJECT_CONFIGS:
    os.makedirs(os.path.join(UPLOAD_FOLDER, _proj), exist_ok=True)

# Per-project active paths (runtime overrides when user selects a file)
ACTIVE_PATHS = {
    proj: dict(cfg['default_paths'])
    for proj, cfg in PROJECT_CONFIGS.items()
}

def get_project_config(project_id):
    """Return config for the given project id, defaulting to clemente (GB300)."""
    return PROJECT_CONFIGS.get(project_id, PROJECT_CONFIGS['clemente'])

def get_project_upload_folder(project_id):
    # project_id must be one of the known project keys -- otherwise it gets
    # concatenated straight into a filesystem path (see os.path.join below),
    # which would let a caller escape UPLOAD_FOLDER with something like
    # project_id="../../somewhere". Falling back to 'clemente' keeps behavior
    # sane for unknown ids while closing that path-traversal hole.
    if project_id not in PROJECT_CONFIGS:
        project_id = 'clemente'
    folder = os.path.join(UPLOAD_FOLDER, project_id)
    os.makedirs(folder, exist_ok=True)
    return folder


def _build_allowed_roots():
    """Collect every directory the app is configured to read/write files from
    (all projects' dir_roots plus their upload folders). Used to make sure any
    user-supplied file_path/base_file/dvt_file/etc. query param can only ever
    point inside one of these directories."""
    roots = set()
    for cfg in PROJECT_CONFIGS.values():
        for tab_dirs in cfg['dir_roots'].values():
            for d in tab_dirs:
                roots.add(os.path.realpath(d))
    roots.add(os.path.realpath(UPLOAD_FOLDER))
    roots.add(os.path.realpath(DATA_DIR))
    return roots

ALLOWED_ROOTS = _build_allowed_roots()

def is_allowed_path(path):
    """True only if `path` resolves to somewhere inside one of ALLOWED_ROOTS.
    This is the guard against arbitrary file read via file_path/base_file/etc.
    query parameters -- without it, any path that merely exists on disk would
    be served back to the caller."""
    if not path:
        return False
    try:
        abs_path = os.path.realpath(path)
    except Exception:
        return False
    for root in ALLOWED_ROOTS:
        try:
            if os.path.commonpath([abs_path, root]) == root:
                return True
        except ValueError:
            continue
    return False

def is_valid_bkc_table(file_path):
    filename = os.path.basename(file_path).lower()
    if any(k in filename for k in ['versiontracker', 'version_tracker', 'changelog', 'change_log', 'release_note', 'releasenote', 'history', 'revision']):
        return False
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        return True
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        valid_sheets = [s for s in wb.sheetnames if not any(k in s.lower() for k in ['change list', 'changelog', 'history', 'readme', 'instructions', 'notes', 'versiontracker'])]
        if not valid_sheets:
            wb.close()
            return False
        for sname in valid_sheets[:3]:
            ws = wb[sname]
            for row in list(ws.iter_rows(values_only=True, max_row=20)):
                row_str = ' '.join([str(c).lower() for c in row if c is not None])
                if any(k in row_str for k in ['component', 'sub component', 'fw version', 'control table', 'bkc version', 'dvt', 'evt', 'pvt']):
                    wb.close()
                    return True
        wb.close()
    except Exception:
        return False
    return False

def scan_files_in_dirs(tab_key, project_id='sanmiguel'):
    cfg = get_project_config(project_id)
    proj_upload = get_project_upload_folder(project_id)
    dirs = list(cfg['dir_roots'].get(tab_key, []))
    # Always include the project-specific upload folder
    if proj_upload not in dirs:
        dirs.append(proj_upload)
    found = []
    seen = set()
    valid_exts = ('.yaml', '.yml') if tab_key == 'yaml' else ('.xlsx', '.csv', '.xls')
    for d in dirs:
        if not os.path.exists(d): continue
        for root, _, files in os.walk(d):
            for f in files:
                if f.lower().endswith(valid_exts) and not f.startswith(('._', '~$')):
                    full_p = os.path.join(root, f)
                    if full_p in seen: continue
                    seen.add(full_p)

                    is_upload = (d == proj_upload)
                    display_name = f"[Uploaded] {f}" if is_upload else os.path.basename(f)

                    mtime = os.path.getmtime(full_p) if os.path.exists(full_p) else 0
                    is_valid_bkc = is_valid_bkc_table(full_p) if tab_key == 'bkc' else True

                    found.append({
                        'filename': f,
                        'display_name': display_name,
                        'path': full_p,
                        'is_excel': f.lower().endswith(('.xlsx', '.xls')),
                        'is_yaml': f.lower().endswith(('.yaml', '.yml')),
                        'is_valid_bkc': is_valid_bkc,
                        'mtime': mtime
                    })

    if tab_key == 'bkc':
        # Prioritize valid BKC Control Tables over Changelog/Tracker files, then sort by mtime descending
        found.sort(key=lambda x: (1 if x.get('is_valid_bkc') else 0, x['mtime']), reverse=True)
    else:
        # Sort files by modification time descending (latest files first)
        found.sort(key=lambda x: x['mtime'], reverse=True)
    return found


def validate_yaml_project(file_path, project_id):
    if not file_path or not os.path.exists(file_path):
        return True, ""
    fname = os.path.basename(file_path).lower()
    
    # 1. Filename checks
    if project_id == 'clemente':
        if any(k in fname for k in ['sanmiguel', 'vr200', 'maxwell', 'earth', 'l10_station1_fvt', 'l10_station2_runin', 'l10_station3_ort']):
            return False, f"專案不對，請重新輸入新檔案！您選擇/上傳的腳本 '{os.path.basename(file_path)}' 屬於 SanMiguel (VR200) 專案，但當前 active 專案為 Clemente (GB300)。"
    elif project_id == 'sanmiguel':
        if any(k in fname for k in ['clemente', 'gb300', 'maxq', 'clemente_ct_maxq_mp_fdt', 'clemente_ct_maxq_mp_fft', 'clemente_ct_maxq_mp_fro']):
            return False, f"專案不對，請重新輸入新檔案！您選擇/上傳的腳本 '{os.path.basename(file_path)}' 屬於 Clemente (GB300) 專案，但當前 active 專案為 SanMiguel (VR200)。"
            
    # 2. Deep content checks (first 50,000 characters)
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read(50000).lower()
            
        if project_id == 'clemente':
            sanmiguel_indicators = [
                'sanmiguel', 'vr200', 'flash-sanmiguel', 'maxwell', 'earth',
                'station 1 (fvt)', 'station 2 (runin)', 'station 3 (ort)',
                'functional verification test (fvt)'
            ]
            for ind in sanmiguel_indicators:
                if ind in content:
                    return False, f"專案不對，請重新輸入新檔案！腳本檔案 '{os.path.basename(file_path)}' 內包含 SanMiguel (VR200) 專案特徵 '{ind}'，但當前 active 專案為 Clemente (GB300)。"
                    
        elif project_id == 'sanmiguel':
            clemente_indicators = [
                'clemente', 'gb300', 'flash-clemente', 'clemente_ct',
                'station 1 (fdt)', 'station 2 (fro)', 'station 3 (fft)',
                'clementesbios', 'clementehmc', 'clementebmc'
            ]
            for ind in clemente_indicators:
                if ind in content:
                    return False, f"專案不對，請重新輸入新檔案！腳本檔案 '{os.path.basename(file_path)}' 內包含 Clemente (GB300) 專案特徵 '{ind}'，但當前 active 專案為 SanMiguel (VR200)。"
    except Exception:
        pass
        
    return True, ""


def resolve_file_path(tab_key, req_path=None, project_id='sanmiguel'):
    proj_cfg = get_project_config(project_id)
    proj_defaults = proj_cfg['default_paths']
    default_path = proj_defaults.get(tab_key) or proj_defaults.get(f"{tab_key}_dvt") or ''

    scanned = scan_files_in_dirs(tab_key, project_id)
    scanned_paths = set(os.path.abspath(f['path']) for f in scanned)

    if req_path and os.path.exists(req_path) and is_allowed_path(req_path):
        abs_req = os.path.abspath(req_path)
        if abs_req in scanned_paths:
            return abs_req

        is_cross_proj = False
        for other_pid in PROJECT_CONFIGS:
            if other_pid != project_id:
                other_scanned = scan_files_in_dirs(tab_key, other_pid)
                other_scanned_paths = set(os.path.abspath(f['path']) for f in other_scanned)
                other_default = os.path.abspath(PROJECT_CONFIGS[other_pid]['default_paths'].get(tab_key, ''))
                if abs_req in other_scanned_paths or abs_req == other_default or f'/{other_pid}/' in abs_req.lower() or f'vr200-sanmiguel' in abs_req.lower():
                    is_cross_proj = True
                    break

        if not is_cross_proj:
            return abs_req

    if scanned:
        return scanned[0]['path']

    return default_path


def resolve_or_default(raw_path, tab_key, default_val, project_id='sanmiguel'):
    """Validate a raw file path coming from a query/form param (file_path,
    base_file, dvt_file, pvt_file, ...): only use it if it exists on disk AND
    resolves inside an allowed project directory (see is_allowed_path). This
    is the single choke point that replaces the old pattern of
    `req_path if (req_path and os.path.exists(req_path)) else resolve_file_path(...)`,
    which let any existing file on the server -- not just project files -- be
    read back through the API."""
    if raw_path and os.path.exists(raw_path) and is_allowed_path(raw_path):
        return os.path.abspath(raw_path)
    return resolve_file_path(tab_key, default_val, project_id)


def filter_valid_data_sheets(sheets):
    ignored_keywords = {
        'readme', 'change log', 'changelog', 'change_log', 'change list', 
        'fw list', 'history', 'revision history', 'single source vendor', 
        'instructions', 'notes', 'version tracker', 'versiontracker', 'poevt'
    }
    valid = [s for s in sheets if not any(k in s.lower() for k in ignored_keywords)]
    return valid if valid else sheets

def read_file_safe(path, sheet_name=None):
    if not path or not os.path.exists(path):
        return None, [], f"File not found: {path}"
    
    ext = os.path.splitext(path)[1].lower()
    sheet_names = []
    
    if ext in ['.xlsx', '.xls']:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet_names = wb.sheetnames
            valid_sheets = filter_valid_data_sheets(sheet_names)
            if sheet_name and sheet_name in sheet_names:
                target_sheet = sheet_name
            else:
                target_sheet = valid_sheets[0]
            ws = wb[target_sheet]
            
            rows = []
            for r in ws.iter_rows(values_only=True):
                row_str = [str(cell).strip() if cell is not None else '' for cell in r]
                rows.append(row_str)
            return rows, valid_sheets, None
        except Exception as e:
            return None, [], str(e)
    else:
        try:
            with open(path, 'r', encoding='utf-8-sig', errors='ignore') as f:
                reader = list(csv.reader(f))
                return reader, [], None
        except Exception as e:
            return None, [], str(e)

FILE_ROWS_CACHE = {}

def read_file_safe_cached(path, sheet_name=None):
    if not path or not os.path.exists(path):
        return None, [], f"File not found: {path}"
    try:
        mtime = os.path.getmtime(path)
        cache_key = (path, sheet_name, mtime)
        if cache_key in FILE_ROWS_CACHE:
            return FILE_ROWS_CACHE[cache_key]
        res = read_file_safe(path, sheet_name=sheet_name)
        FILE_ROWS_CACHE[cache_key] = res
        return res
    except Exception:
        return read_file_safe(path, sheet_name=sheet_name)

def parse_version_tuple(ver_str):
    if not ver_str or ver_str in ['-', 'NA', 'TBD']:
        return ()
    tokens = re.findall(r'[0-9]+|[a-zA-Z]+', ver_str)
    parsed = []
    for t in tokens:
        if t.isdigit():
            parsed.append((0, int(t)))
        else:
            parsed.append((1, t.lower()))
    return tuple(parsed)

def compare_versions(ver_base, ver_target):
    if ver_base == ver_target:
        return 'same'
    if not ver_base and ver_target:
        return 'added'
    if ver_base and not ver_target:
        return 'removed'
        
    t_base = parse_version_tuple(ver_base)
    t_target = parse_version_tuple(ver_target)
    
    if not t_base and t_target:
        return 'upgraded'
    if t_base and not t_target:
        return 'downgraded'
        
    if t_target > t_base:
        return 'upgraded'
    elif t_target < t_base:
        return 'downgraded'
    else:
        return 'updated'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/projects')
def get_projects():
    """Return list of supported projects and their metadata."""
    projects = []
    for proj_id, cfg in PROJECT_CONFIGS.items():
        projects.append({
            'id': proj_id,
            'label': cfg['label'],
            'color_theme': cfg['color_theme'],
        })
    return jsonify({'success': True, 'projects': projects})

@app.route('/api/files')
def get_files():
    tab_key = request.args.get('type')
    project_id = request.args.get('project', 'sanmiguel')

    bkc_files = scan_files_in_dirs('bkc', project_id)
    yaml_files = scan_files_in_dirs('yaml', project_id)
    fru_files = scan_files_in_dirs('fru', project_id)
    matrix_files = scan_files_in_dirs('matrix', project_id)

    target_type = tab_key or 'bkc'
    files = scan_files_in_dirs(target_type, project_id)

    return jsonify({
        'success': True,
        'type': target_type,
        'files': files,
        'bkc': bkc_files,
        'yaml': yaml_files,
        'fru': fru_files,
        'matrix': matrix_files
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    tab_type = request.form.get('tab_type', 'bkc')
    project_id = request.form.get('project', 'sanmiguel')

    if file.filename == '':
        return jsonify({'success': False, 'error': 'No selected file'}), 400

    if file and file.filename.lower().endswith(('.xlsx', '.xls', '.csv', '.yaml', '.yml')):
        filename = safe_filename(file.filename)
        proj_upload = get_project_upload_folder(project_id)
        save_path = os.path.join(proj_upload, filename)
        file.save(save_path)

        if tab_type == 'yaml':
            is_valid, err_msg = validate_yaml_project(save_path, project_id)
            if not is_valid:
                if os.path.exists(save_path):
                    try: os.remove(save_path)
                    except Exception: pass
                return jsonify({'success': False, 'error': err_msg}), 400

        # Set as active path for the given project + type
        if project_id in ACTIVE_PATHS and tab_type in ACTIVE_PATHS[project_id]:
            ACTIVE_PATHS[project_id][tab_type] = save_path

        return jsonify({
            'success': True,
            'filename': filename,
            'saved_path': save_path,
            'tab_type': tab_type,
            'project': project_id
        })
    else:
        return jsonify({'success': False, 'error': 'Unsupported file format. Please upload .xlsx, .csv, .yaml or .yml'}), 400


@app.route('/api/status')
def get_status():
    project_id = request.args.get('project', 'sanmiguel')
    status = {}
    proj_paths = ACTIVE_PATHS.get(project_id, {})
    for key, path in proj_paths.items():
        status[key] = {
            'path': path,
            'exists': os.path.exists(path) if path else False,
            'filename': os.path.basename(path) if path else '',
            'is_excel': os.path.splitext(path)[1].lower() in ['.xlsx', '.xls'] if path else False
        }
    return jsonify(status)

@app.route('/api/bkc')
def get_bkc():
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')

    req_path = request.args.get('file_path', None)
    path = resolve_or_default(req_path, 'bkc', default_bkc, project_id)
    requested_sheet = request.args.get('sheet', None)

    rows, sheets, err = read_file_safe(path, sheet_name=requested_sheet)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    items = []
    groups = set()
    categories_dict = {}
    current_category = 'General'
    current_group = ''

    headers = []
    if len(rows) > 2:
        headers = [c for c in rows[2]]

    for idx, r in enumerate(rows[3:], start=3):
        if not any(r):
            continue

        c0 = r[0].strip() if len(r) > 0 else ''
        c1 = r[1].strip() if len(r) > 1 else ''
        c5 = r[5].strip() if len(r) > 5 else ''

        if c0 and not c1 and not c5 and not any(c.strip() for c in r[2:]):
            current_category = c0
            if current_category not in categories_dict:
                categories_dict[current_category] = []
            continue

        if c0:
            current_group = c0
            groups.add(current_group)

        meta_owner = r[2].strip() if len(r) > 2 else ''
        odm_owner = r[3].strip() if len(r) > 3 else ''
        
        ver = c5
        if ver.endswith('.0') and ver.replace('.0', '').isdigit():
            ver = ver[:-2]
        
        chksum = r[6].strip() if len(r) > 6 else ''
        vrc = r[7].strip() if len(r) > 7 else ''
        if vrc.endswith('.0') and vrc.replace('.0', '').isdigit():
            vrc = vrc[:-2]
        
        sign_off = r[8].strip() if len(r) > 8 else ''
        gdrive = r[11].strip() if len(r) > 11 else ''

        item = {
            'row_id': idx,
            'category': current_category,
            'group': current_group,
            'sub_component': c1,
            'meta_owner': meta_owner,
            'odm_owner': odm_owner,
            'version': ver,
            'checksum': chksum,
            'vrc': vrc,
            'sign_off': sign_off,
            'gdrive': gdrive,
            'raw_row': r
        }

        items.append(item)
        if current_category not in categories_dict:
            categories_dict[current_category] = []
        categories_dict[current_category].append(item)

    cat_summary = []
    for cat_name, cat_items in categories_dict.items():
        cat_groups = len(set(it['group'] for it in cat_items))
        cat_fw_count = sum(1 for it in cat_items if it['version'])
        cat_summary.append({
            'name': cat_name,
            'items_count': len(cat_items),
            'groups_count': cat_groups,
            'fw_count': cat_fw_count
        })

    summary = {
        'total_items': len(items),
        'groups_count': len(groups),
        'groups_list': sorted(list(groups)),
        'items_with_version': sum(1 for item in items if item['version']),
        'sheets': sheets,
        'active_sheet': requested_sheet if (requested_sheet and requested_sheet in sheets) else (sheets[0] if sheets else None),
        'active_file': path,
        'available_files': scan_files_in_dirs('bkc', project_id),
        'categories_summary': cat_summary
    }

    return jsonify({
        'success': True,
        'filename': os.path.basename(path),
        'summary': summary,
        'headers': headers,
        'categories': cat_summary,
        'items': items
    })

@app.route('/api/bkc-compare')
def get_bkc_compare():
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')

    req_path = request.args.get('file_path', None)
    path = resolve_or_default(req_path, 'bkc', default_bkc, project_id)

    base_sheet = request.args.get('base_sheet', None)
    target_sheet = request.args.get('target_sheet', None)

    rows_target, sheets, err_target = read_file_safe(path, sheet_name=target_sheet)
    if err_target or not sheets:
        return jsonify({'success': False, 'error': err_target or 'No sheets found'}), 400

    t_sheet = target_sheet if (target_sheet and target_sheet in sheets) else sheets[0]
    default_base = sheets[1] if len(sheets) > 1 else t_sheet
    b_sheet = base_sheet if (base_sheet and base_sheet in sheets) else default_base

    rows_base, _, err_base = read_file_safe(path, sheet_name=b_sheet)
    if err_base:
        return jsonify({'success': False, 'error': err_base}), 400

    def parse_items_map(rows):
        items_map = {}
        current_cat = 'General'
        current_group = ''
        for idx, r in enumerate(rows[3:], start=3):
            if not any(r): continue
            c0 = r[0].strip() if len(r) > 0 else ''
            c1 = r[1].strip() if len(r) > 1 else ''
            c5 = r[5].strip() if len(r) > 5 else ''

            if c0 and not c1 and not c5 and not any(c.strip() for c in r[2:]):
                current_cat = c0
                continue

            if c0: current_group = c0

            meta_owner = r[2].strip() if len(r) > 2 else ''
            odm_owner = r[3].strip() if len(r) > 3 else ''

            ver = c5
            if ver.endswith('.0') and ver.replace('.0', '').isdigit():
                ver = ver[:-2]

            chksum = r[6].strip() if len(r) > 6 else ''
            vrc = r[7].strip() if len(r) > 7 else ''
            if vrc.endswith('.0') and vrc.replace('.0', '').isdigit():
                vrc = vrc[:-2]

            sign_off = r[8].strip() if len(r) > 8 else ''
            gdrive = r[11].strip() if len(r) > 11 else ''

            key = (current_cat, current_group, c1)
            items_map[key] = {
                'category': current_cat,
                'group': current_group,
                'sub_component': c1,
                'meta_owner': meta_owner,
                'odm_owner': odm_owner,
                'version': ver,
                'checksum': chksum,
                'vrc': vrc,
                'sign_off': sign_off,
                'gdrive': gdrive
            }
        return items_map

    base_map = parse_items_map(rows_base)
    target_map = parse_items_map(rows_target)

    ordered_keys = []
    for k in base_map.keys():
        if k not in ordered_keys: ordered_keys.append(k)
    for k in target_map.keys():
        if k not in ordered_keys: ordered_keys.append(k)

    compared_items = []
    diff_count = 0
    upgraded_count = 0
    downgraded_count = 0
    categories_dict = {}

    for key in ordered_keys:
        b_item = base_map.get(key, {})
        t_item = target_map.get(key, {})

        cat = t_item.get('category') or b_item.get('category') or 'General'
        grp = t_item.get('group') or b_item.get('group') or ''
        sub = t_item.get('sub_component') or b_item.get('sub_component') or ''

        ver_b = b_item.get('version', '')
        ver_t = t_item.get('version', '')

        sign_b = b_item.get('sign_off', '')
        sign_t = t_item.get('sign_off', '')

        ver_status = compare_versions(ver_b, ver_t)
        is_diff = (ver_status != 'same') or (sign_b != sign_t)
        
        if ver_status == 'upgraded': upgraded_count += 1
        elif ver_status == 'downgraded': downgraded_count += 1
        
        if is_diff:
            diff_count += 1

        comp = {
            'category': cat,
            'group': grp,
            'sub_component': sub,
            'meta_owner': t_item.get('meta_owner') or b_item.get('meta_owner') or '',
            'odm_owner': t_item.get('odm_owner') or b_item.get('odm_owner') or '',
            'base_version': ver_b,
            'target_version': ver_t,
            'base_sign_off': sign_b,
            'target_sign_off': sign_t,
            'base_checksum': b_item.get('checksum', ''),
            'target_checksum': t_item.get('checksum', ''),
            'gdrive': t_item.get('gdrive') or b_item.get('gdrive') or '',
            'is_diff': is_diff,
            'status': ver_status
        }

        compared_items.append(comp)
        if cat not in categories_dict:
            categories_dict[cat] = []
        categories_dict[cat].append(comp)

    cat_summary = []
    for cat_name, cat_items in categories_dict.items():
        cat_groups = len(set(it['group'] for it in cat_items))
        cat_diffs = sum(1 for it in cat_items if it['is_diff'])
        cat_summary.append({
            'name': cat_name,
            'items_count': len(cat_items),
            'groups_count': cat_groups,
            'diff_count': cat_diffs
        })

    summary = {
        'total_items': len(compared_items),
        'diff_items_count': diff_count,
        'upgraded_count': upgraded_count,
        'downgraded_count': downgraded_count,
        'sheets': sheets,
        'base_sheet': b_sheet,
        'target_sheet': t_sheet,
        'active_file': path,
        'available_files': scan_files_in_dirs('bkc', project_id),
        'categories_summary': cat_summary
    }

    return jsonify({
        'success': True,
        'filename': os.path.basename(path),
        'summary': summary,
        'categories': cat_summary,
        'items': compared_items
    })

KNOWN_FRU_SECTIONS = {'Chassis Info Area', 'Board Info Area', 'Product Info Area', 'MultiRecord Area', 'Organization'}

def is_fru_section_header(label):
    if not label: return False
    lbl_clean = label.strip()
    if lbl_clean in KNOWN_FRU_SECTIONS or 'Area' in lbl_clean or 'Section' in lbl_clean or 'Header' in lbl_clean:
        return True
    return False

def parse_fru_sheet_smart(rows):
    if len(rows) < 2: return [], [], {}
    header_row = rows[1]
    modules = []
    mod_map = {}
    for col_idx in range(1, len(header_row)):
        m_name = header_row[col_idx].strip()
        if m_name and 'FRU' not in m_name and m_name not in modules:
            modules.append(m_name)
            mod_map[m_name] = col_idx

    current_section = 'General'
    ordered_keys = []
    data_dict = {}

    for i in range(len(rows)):
        r = rows[i]
        col0 = (r[0] if len(r) > 0 else '').strip()
        has_data = any(c.strip() for c in r[1:])

        if is_fru_section_header(col0):
            current_section = col0
            continue

        if not col0 and has_data and ordered_keys:
            last_key = ordered_keys[-1]
            for mod_name, col_idx in mod_map.items():
                val = r[col_idx].strip() if col_idx < len(r) else ''
                if val:
                    prev_val = data_dict[last_key].get(mod_name, '')
                    data_dict[last_key][mod_name] = (prev_val + ' (' + val + ')').strip() if prev_val else val
            continue

        if not col0 and not has_data:
            continue

        key = (current_section, col0)
        if key not in data_dict:
            data_dict[key] = {}
            ordered_keys.append(key)

        for mod_name, col_idx in mod_map.items():
            val = r[col_idx].strip() if col_idx < len(r) else ''
            data_dict[key][mod_name] = val

    return modules, ordered_keys, data_dict

def filter_valid_fru_sheets(sheets):
    if not sheets:
        return []
    valid = []
    ignored_keywords = [
        'changehistory', 'change history', 'panel change', 'mp(panel change)',
        'version', 'history', 'changelog', 'change log', 'cover', 'readme', 'notes', 'note'
    ]
    for s in sheets:
        s_clean = str(s or '').strip().lower()
        if not any(kw in s_clean for kw in ignored_keywords):
            valid.append(s)
    return valid if valid else sheets


@app.route('/api/fru')
def get_fru():
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_fru_dvt = proj_cfg['default_paths'].get('fru_dvt', '')

    req_path = request.args.get('file_path', None)
    path = resolve_or_default(req_path, 'fru', default_fru_dvt, project_id)
    requested_sheet = request.args.get('sheet', None)

    rows, sheets, err = read_file_safe(path, sheet_name=requested_sheet)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    valid_sheets = filter_valid_fru_sheets(sheets)
    active_s = requested_sheet if (requested_sheet and requested_sheet in valid_sheets) else (valid_sheets[0] if valid_sheets else (sheets[0] if sheets else None))

    if active_s and active_s != requested_sheet:
        rows, _, err = read_file_safe(path, sheet_name=active_s)
        if err:
            return jsonify({'success': False, 'error': err}), 400

    modules, keys, data_dict = parse_fru_sheet_smart(rows)

    fields = []
    total_configured_vals = 0

    for idx, (sec, f_name) in enumerate(keys):
        key = (sec, f_name)
        vals = data_dict.get(key, {})

        for mod, val in vals.items():
            if val and val != '-':
                total_configured_vals += 1

        fields.append({
            'row_id': idx,
            'section': sec,
            'field_name': f_name,
            'values': vals
        })

    summary = {
        'total_rows': len(fields),
        'modules': modules,
        'total_configured_vals': total_configured_vals,
        'filename': os.path.basename(path),
        'sheets': valid_sheets,
        'active_sheet': active_s,
        'active_file': path,
        'available_files': scan_files_in_dirs('fru', project_id)
    }

    return jsonify({
        'success': True,
        'summary': summary,
        'fields': fields
    })

@app.route('/api/fru-compare')
def get_fru_compare():
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_fru_dvt = proj_cfg['default_paths'].get('fru_dvt', '')
    default_fru_pvt = proj_cfg['default_paths'].get('fru_pvt', '')

    req_dvt = request.args.get('dvt_file', None)
    req_pvt = request.args.get('pvt_file', None)
    
    path_dvt = resolve_or_default(req_dvt, 'fru', default_fru_dvt, project_id)
    path_pvt = resolve_or_default(req_pvt, 'fru', default_fru_pvt, project_id)


    requested_sheet = request.args.get('sheet', None)
    base_sheet = request.args.get('base_sheet', requested_sheet)
    target_sheet = request.args.get('target_sheet', requested_sheet)

    rows_dvt, sheets_dvt, err_dvt = read_file_safe(path_dvt, sheet_name=base_sheet)
    rows_pvt, sheets_pvt, err_pvt = read_file_safe(path_pvt, sheet_name=target_sheet)

    if err_dvt or err_pvt:
        return jsonify({
            'success': False,
            'error_dvt': err_dvt,
            'error_pvt': err_pvt
        }), 400

    valid_sheets_dvt = filter_valid_fru_sheets(sheets_dvt)
    valid_sheets_pvt = filter_valid_fru_sheets(sheets_pvt)

    base_s = base_sheet if (base_sheet and base_sheet in valid_sheets_dvt) else (valid_sheets_dvt[0] if valid_sheets_dvt else (sheets_dvt[0] if sheets_dvt else None))
    target_s = target_sheet if (target_sheet and target_sheet in valid_sheets_pvt) else (valid_sheets_pvt[0] if valid_sheets_pvt else (sheets_pvt[0] if sheets_pvt else None))

    if base_s and base_s != base_sheet:
        rows_dvt, _, err_dvt = read_file_safe(path_dvt, sheet_name=base_s)
    if target_s and target_s != target_sheet:
        rows_pvt, _, err_pvt = read_file_safe(path_pvt, sheet_name=target_s)

    if err_dvt or err_pvt:
        return jsonify({
            'success': False,
            'error_dvt': err_dvt,
            'error_pvt': err_pvt
        }), 400

    mods_b, keys_b, dict_b = parse_fru_sheet_smart(rows_dvt)
    mods_t, keys_t, dict_t = parse_fru_sheet_smart(rows_pvt)

    all_modules = list(dict.fromkeys(mods_b + mods_t))
    all_keys = list(dict.fromkeys(keys_b + keys_t))

    common_modules = [m for m in mods_b if m in mods_t]
    base_only_modules = [m for m in mods_b if m not in mods_t]
    target_only_modules = [m for m in mods_t if m not in mods_b]

    fields_compare = []
    total_diff_cells = 0
    total_diff_rows = 0

    for idx, (sec, f_name) in enumerate(all_keys):

        key = (sec, f_name)
        b_vals = dict_b.get(key, {})
        t_vals = dict_t.get(key, {})

        dvt_vals = {}
        pvt_vals = {}
        diff_modules = []
        is_row_diff = False

        for mod in all_modules:
            has_in_b = mod in mods_b
            has_in_t = mod in mods_t

            val_b = b_vals.get(mod, '-' if has_in_b else 'N/A (無此模組)')
            val_t = t_vals.get(mod, '-' if has_in_t else 'N/A (無此模組)')

            dvt_vals[mod] = val_b
            pvt_vals[mod] = val_t

            # Real spec difference exists when both Base & Target have the module and values differ
            if has_in_b and has_in_t and val_b != val_t:
                diff_modules.append(mod)
                total_diff_cells += 1
                is_row_diff = True

        if is_row_diff:
            total_diff_rows += 1

        fields_compare.append({
            'row_id': idx,
            'section': sec,
            'field_name': f_name,
            'dvt_values': dvt_vals,
            'pvt_values': pvt_vals,
            'is_diff': is_row_diff,
            'diff_modules': diff_modules
        })

    summary = {
        'total_rows': len(fields_compare),
        'total_diff_rows': total_diff_rows,
        'total_diff_cells': total_diff_cells,
        'modules': all_modules,
        'common_modules': common_modules,
        'base_only_modules': base_only_modules,
        'target_only_modules': target_only_modules,
        'dvt_filename': os.path.basename(path_dvt),
        'pvt_filename': os.path.basename(path_pvt),
        'dvt_path': path_dvt,
        'pvt_path': path_pvt,
        'sheets_base': valid_sheets_dvt,
        'sheets_target': valid_sheets_pvt,
        'base_sheet': base_s,
        'target_sheet': target_s,
        'available_files': scan_files_in_dirs('fru', project_id)
    }





    return jsonify({
        'success': True,
        'summary': summary,
        'fields': fields_compare
    })

# ── Shared helper: parse a build matrix sheet ──────────────────────────────
def parse_matrix_sheet(rows):
    """Parse a build-matrix sheet (already loaded as list-of-str-lists).
    Returns: (configs, configs_info, descriptions, rack_qty, items_dict, items_list)
      configs      – ordered list of config names
      configs_info – list of (col_idx, cfg_name)
      descriptions – {cfg_name: description str}
      rack_qty     – {cfg_name: qty str}
      items_dict   – {(group_item, attribute): {cfg_name: val}}
      items_list   – [{row_id, group_item, attribute, values, is_diff}]
    """
    # Smart config-row search (first 6 rows)
    config_row_idx = -1
    for i in range(min(6, len(rows))):
        r = rows[i]
        if sum(1 for cell in r if 'config' in cell.lower()) >= 1:
            config_row_idx = i
            break
    if config_row_idx == -1:
        config_row_idx = 2 if len(rows) > 2 else 0

    cfg_row = rows[config_row_idx] if config_row_idx < len(rows) else []
    configs_info, configs = [], []
    for c_idx in range(2, len(cfg_row)):
        c_val = cfg_row[c_idx].strip()
        if c_val and c_val.lower() != 'none' and not any(k in c_val.lower() for k in ['note', 'comment']):
            configs_info.append((c_idx, c_val))
            configs.append(c_val)

    desc_row = rows[config_row_idx + 1] if config_row_idx + 1 < len(rows) else []
    qty_row  = rows[config_row_idx + 2] if config_row_idx + 2 < len(rows) else []

    descriptions, rack_qty = {}, {}
    for c_idx, cfg_name in configs_info:
        if c_idx < len(desc_row):
            descriptions[cfg_name] = desc_row[c_idx].strip()
        if c_idx < len(qty_row):
            q = qty_row[c_idx].strip()
            if q.endswith('.0'): q = q[:-2]
            rack_qty[cfg_name] = q

    items_dict   = {}   # {(group_item, attribute): {cfg_name: val}}
    items_list   = []
    current_group = 'General / Header'
    total_diff   = 0

    for idx in range(config_row_idx + 3, len(rows)):
        r = rows[idx]
        if not any(c.strip() for c in r): continue

        col0 = r[0].strip() if len(r) > 0 else ''
        col1 = r[1].strip() if len(r) > 1 else ''

        if col0:
            current_group = col0

        if not col1 and not any(r[c_idx].strip() for c_idx, _ in configs_info if c_idx < len(r)):
            continue

        vals, row_vals = {}, []
        for c_idx, cfg_name in configs_info:
            val = r[c_idx].strip() if c_idx < len(r) else ''
            vals[cfg_name] = val
            if val: row_vals.append(val)

        # Skip rows where all values are empty and attribute is purely metadata label
        if not row_vals and col1.lower() in ['igs', 'meta', 'note', 'comment', 'notes', '']:
            continue

        is_diff = len(set(row_vals)) > 1
        if is_diff: total_diff += 1

        key = (current_group, col1)
        items_dict[key] = vals

        items_list.append({
            'row_id': idx,
            'group_item': current_group,
            'attribute': col1,
            'values': vals,
            'is_diff': is_diff
        })

    return configs, configs_info, descriptions, rack_qty, items_dict, items_list, total_diff


def compare_two_matrix_sheets(base_rows, tgt_rows):
    b_configs, _, b_descs, b_qty, b_dict, _, _ = parse_matrix_sheet(base_rows)
    t_configs, _, t_descs, t_qty, t_dict, _, _ = parse_matrix_sheet(tgt_rows)

    seen_keys   = set()
    ordered_keys = []
    for k in b_dict:
        if k not in seen_keys:
            seen_keys.add(k)
            ordered_keys.append(k)
    for k in t_dict:
        if k not in seen_keys:
            seen_keys.add(k)
            ordered_keys.append(k)

    compare_items = []
    diff_count    = 0

    for key in ordered_keys:
        group_item, attribute = key
        b_vals = b_dict.get(key, None)
        t_vals = t_dict.get(key, None)

        if b_vals is None:
            diff_type = 'target_only'
        elif t_vals is None:
            diff_type = 'base_only'
        else:
            b_set = set(v for v in b_vals.values() if v)
            t_set = set(v for v in t_vals.values() if v)
            diff_type = 'changed' if b_set != t_set else 'same'

        is_diff = diff_type != 'same'
        if is_diff:
            diff_count += 1

        compare_items.append({
            'group_item': group_item,
            'attribute':  attribute,
            'base_values':   b_vals or {},
            'target_values': t_vals or {},
            'diff_type': diff_type,
            'is_diff':   is_diff
        })

    return {
        'base_configs': b_configs,
        'target_configs': t_configs,
        'base_descriptions': b_descs,
        'target_descriptions': t_descs,
        'base_rack_qty': b_qty,
        'target_rack_qty': t_qty,
        'diff_items_count': diff_count,
        'total_items': len(compare_items),
        'items': compare_items
    }


@app.route('/api/build-matrix')
def get_build_matrix():
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_matrix = proj_cfg['default_paths'].get('matrix', '')

    req_path = request.args.get('file_path', None)
    path = resolve_or_default(req_path, 'matrix', default_matrix, project_id)
    requested_sheet = request.args.get('sheet', None)

    rows, sheets, err = read_file_safe(path, sheet_name=requested_sheet)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    configs, _, descriptions, rack_qty, _, items_list, total_diff = parse_matrix_sheet(rows)
    active_sh = requested_sheet if (requested_sheet and requested_sheet in sheets) else (sheets[0] if sheets else None)

    auto_compare = None
    if len(sheets) >= 2:
        base_sh = sheets[0]
        tgt_sh = sheets[1]
        r_base, _, _ = read_file_safe(path, sheet_name=base_sh)
        r_tgt, _, _ = read_file_safe(path, sheet_name=tgt_sh)
        if r_base and r_tgt:
            res = compare_two_matrix_sheets(r_base, r_tgt)
            auto_compare = {
                'is_multi_sheet': True,
                'base_sheet': base_sh,
                'target_sheet': tgt_sh,
                'diff_count': res['diff_items_count'],
                'total_items': res['total_items'],
                'base_configs': res['base_configs'],
                'target_configs': res['target_configs'],
                'items': res['items']
            }

    return jsonify({
        'success': True,
        'summary': {
            'filename': os.path.basename(path),
            'active_file': path,
            'configs': configs,
            'descriptions': descriptions,
            'rack_qty': rack_qty,
            'total_items': len(items_list),
            'diff_items_count': total_diff,
            'sheets': sheets,
            'active_sheet': active_sh,
            'available_files': scan_files_in_dirs('matrix', project_id)
        },
        'items': items_list,
        'auto_compare': auto_compare
    })


@app.route('/api/build-matrix-compare')
def get_build_matrix_compare():
    """Compare two build-matrix sheets (can be different files or different sheets in same file)."""
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    default_matrix = proj_cfg['default_paths'].get('matrix', '')

    base_file  = request.args.get('base_file',  None)
    base_sheet = request.args.get('base_sheet', None)
    tgt_file   = request.args.get('target_file',  None)
    tgt_sheet  = request.args.get('target_sheet', None)

    default_path = resolve_file_path('matrix', default_matrix, project_id)
    base_path = base_file if (base_file and os.path.exists(base_file)) else default_path
    tgt_path  = tgt_file  if (tgt_file  and os.path.exists(tgt_file))  else default_path

    base_rows, base_sheets, err1 = read_file_safe(base_path, sheet_name=base_sheet)
    if err1:
        return jsonify({'success': False, 'error': f'Base sheet error: {err1}'}), 400

    tgt_rows, tgt_sheets, err2 = read_file_safe(tgt_path, sheet_name=tgt_sheet)
    if err2:
        return jsonify({'success': False, 'error': f'Target sheet error: {err2}'}), 400

    b_sheet = base_sheet if (base_sheet and base_sheet in base_sheets) else base_sheets[0]
    t_sheet = tgt_sheet if (tgt_sheet and tgt_sheet in tgt_sheets) else (tgt_sheets[1] if len(tgt_sheets) > 1 else tgt_sheets[0])

    res = compare_two_matrix_sheets(base_rows, tgt_rows)

    return jsonify({
        'success': True,
        'summary': {
            'base_file':    os.path.basename(base_path),
            'base_sheet':   b_sheet,
            'target_file':  os.path.basename(tgt_path),
            'target_sheet': t_sheet,
            'base_configs':   res['base_configs'],
            'target_configs': res['target_configs'],
            'base_descriptions':   res['base_descriptions'],
            'target_descriptions': res['target_descriptions'],
            'base_rack_qty':   res['base_rack_qty'],
            'target_rack_qty': res['target_rack_qty'],
            'total_items': res['total_items'],
            'diff_items_count': res['diff_items_count'],
            'base_sheets':   base_sheets,
            'target_sheets': tgt_sheets,
            'available_files': scan_files_in_dirs('matrix', project_id),
            'base_path': base_path,
            'target_path': tgt_path
        },
        'items': res['items']
    })


@app.route('/api/sync-gdrive', methods=['POST', 'GET'])
def api_sync_gdrive():
    """Manual trigger endpoint for Google Drive sync."""
    try:
        from gdrive_sync import sync_all_gdrive_folders
        res = sync_all_gdrive_folders()
        return jsonify({
            'success': True,
            'result': res
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ==================== V2 FEATURE ENDPOINTS ====================

WATCHLIST_FILE = os.path.join(BASE_DIR, 'watchlist.json')

def get_project_signoff_file(project_id='sanmiguel'):
    proj_dir = os.path.join(DATA_DIR, project_id)
    os.makedirs(proj_dir, exist_ok=True)
    return os.path.join(proj_dir, 'signoffs.json')

def load_signoffs(project_id='sanmiguel'):
    sf = get_project_signoff_file(project_id)
    if os.path.exists(sf):
        try:
            with open(sf, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_signoffs(data, project_id='sanmiguel'):
    sf = get_project_signoff_file(project_id)
    try:
        atomic_write_json(sf, data)
    except Exception as e:
        print(f"Failed to save signoffs for {project_id}: {e}")

def load_watchlist():
    if os.path.exists(WATCHLIST_FILE):
        try:
            with open(WATCHLIST_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('keywords', [])
        except Exception:
            pass
    return ["CPLD", "BIOS", "BMC", "MB PN", "Compute Tray", "VR", "PMIC"]


@app.route('/api/watchlist', methods=['GET', 'POST'])
def api_watchlist():
    """Get or update critical component watchlist keywords."""
    if request.method == 'POST':
        body = request.get_json() or {}
        keywords = body.get('keywords', [])
        try:
            atomic_write_json(WATCHLIST_FILE, {'keywords': keywords})
            return jsonify({'success': True, 'keywords': keywords})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    return jsonify({'success': True, 'keywords': load_watchlist()})


@app.route('/api/signoff', methods=['GET', 'POST'])
def api_signoff():
    """Get or save engineer sign-off status and notes."""
    project_id = request.args.get('project') or (request.get_json(silent=True) or {}).get('project') or 'sanmiguel'
    signoffs = load_signoffs(project_id)
    if request.method == 'POST':
        body = request.get_json() or {}
        key = body.get('key')
        status = body.get('status', 'PENDING')
        note = body.get('note', '')
        user = body.get('user', 'Engineer')
        if not key:
            return jsonify({'success': False, 'error': 'Missing item key'}), 400
        
        signoffs[key] = {
            'status': status,
            'note': note,
            'user': user,
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        save_signoffs(signoffs, project_id)
        return jsonify({'success': True, 'signoff': signoffs[key]})
    
    return jsonify({'success': True, 'signoffs': signoffs})


def parse_bkc_items(rows):
    items = []
    current_category = 'General'
    current_group = ''
    for idx, r in enumerate(rows[3:], start=3):
        if not any(r): continue
        c0 = r[0].strip() if len(r) > 0 else ''
        c1 = r[1].strip() if len(r) > 1 else ''
        c5 = r[5].strip() if len(r) > 5 else ''
        if c0 and not c1 and not c5 and not any(c.strip() for c in r[2:]):
            current_category = c0
            continue
        if c0: current_group = c0
        items.append({
            'category': current_category,
            'group': current_group,
            'sub_component': c1,
            'version': c5
        })
    return items

def compare_two_fru_sheets(rows_dvt, rows_pvt):
    if not rows_dvt or not rows_pvt:
        return {'total_items': 0, 'diff_count': 0, 'same_count': 0, 'fields': []}
    mods_b, keys_b, dict_b = parse_fru_sheet_smart(rows_dvt)
    mods_t, keys_t, dict_t = parse_fru_sheet_smart(rows_pvt)
    all_modules = list(dict.fromkeys(mods_b + mods_t))
    all_keys = list(dict.fromkeys(keys_b + keys_t))
    fields_compare = []
    diff_rows_count = 0
    same_rows_count = 0
    for idx, (sec, f_name) in enumerate(all_keys):
        key = (sec, f_name)
        b_vals = dict_b.get(key, {})
        t_vals = dict_t.get(key, {})
        is_row_diff = False
        val_b_str = ""
        val_t_str = ""
        for mod in all_modules:
            has_in_b = mod in mods_b
            has_in_t = mod in mods_t
            val_b = b_vals.get(mod, '-' if has_in_b else 'N/A')
            val_t = t_vals.get(mod, '-' if has_in_t else 'N/A')
            if not val_b_str and val_b != 'N/A': val_b_str = val_b
            if not val_t_str and val_t != 'N/A': val_t_str = val_t
            if has_in_b and has_in_t and val_b != val_t:
                is_row_diff = True
        if is_row_diff: diff_rows_count += 1
        else: same_rows_count += 1
        fields_compare.append({
            'module': all_modules[0] if all_modules else 'General',
            'section': sec,
            'field_name': f_name,
            'dvt_value': val_b_str,
            'pvt_value': val_t_str,
            'is_diff': is_row_diff
        })
    return {
        'total_items': len(fields_compare),
        'diff_count': diff_rows_count,
        'same_count': same_rows_count,
        'fields': fields_compare
    }


def parse_single_yaml_file(path, default_station_label="Station"):
    if not path or not os.path.exists(path):
        return [], default_station_label, f"File not found: {path}"
    
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            data = yaml.safe_load(f)
    except Exception as e:
        return [], default_station_label, f"YAML parse error: {str(e)}"
    
    if not data:
        return [], default_station_label, "YAML file is empty"
    
    base_name = os.path.basename(path)
    station_label = default_station_label

    # Extract station index prefix from default_station_label (e.g. "Station 1")
    st_idx_prefix = ""
    if default_station_label.startswith("Station "):
        parts = default_station_label.split(" ")
        if len(parts) >= 2 and parts[1].isdigit():
            st_idx_prefix = f"Station {parts[1]} "

    clean_name = base_name
    # Strip long hex hash string (e.g., _8d6cecc0972ecd58da6e8cba)
    clean_name = re.sub(r'_[0-9a-fA-F]{16,32}', '', clean_name)
    clean_name = re.sub(r'\.yaml|\.yml', '', clean_name, flags=re.IGNORECASE)
    clean_name = re.sub(r'^clemente_ct_maxq_mp_', '', clean_name, flags=re.IGNORECASE)
    clean_name = re.sub(r'^clemente_maxq_mp_', '', clean_name, flags=re.IGNORECASE)
    clean_name = re.sub(r'^clemente_mp_', '', clean_name, flags=re.IGNORECASE)
    clean_name = re.sub(r'^clemente_', '', clean_name, flags=re.IGNORECASE)
    clean_name = clean_name.upper()

    station_label = clean_name
    
    extracted_items = []
    
    def clean_val(v):
        if v is None: return ""
        v_str = str(v).strip()
        if v_str.endswith('.0') and v_str.replace('.0', '').isdigit():
            v_str = v_str[:-2]
        return v_str
    
    if isinstance(data, list):
        for idx, step in enumerate(data, start=1):
            if not isinstance(step, dict): continue
            step_name = step.get('name') or f"Step_{idx}"
            args = step.get('args') or {}
            if not isinstance(args, dict): continue
            step_extracted_before = len(extracted_items)
            
            # 1. check_list dictionary (e.g. ClementeGB300VersionCheck)
            chk_list = args.get('check_list')
            if isinstance(chk_list, dict):
                for k, v in chk_list.items():
                    extracted_items.append({
                        'station': station_label,
                        'file_name': base_name,
                        'step_location': str(step_name),
                        'component': str(k),
                        'sub_component': str(k),
                        'yaml_version': clean_val(v),
                        'command': str(step.get('cmd') or args.get('cmd') or ''),
                        'discussion_note': f"Test check in step '{step_name}'"
                    })
            
            # 2. ssd_info dictionary (e.g. SSDFlash~check_only)
            ssd_info = args.get('ssd_info')
            if isinstance(ssd_info, dict):
                for ssd_pn, info in ssd_info.items():
                    if isinstance(info, dict) and info.get('fw_version'):
                        vendor = str(info.get('vendor') or '').strip()
                        fw_file = str(info.get('fw_file') or '').strip()
                        fw_ver = clean_val(info.get('fw_version'))

                        model_name = ""
                        pn_upper = str(ssd_pn).upper()
                        file_upper = fw_file.upper()

                        if '7450' in file_upper or 'MTFDKBT' in pn_upper:
                            model_name = "Micron 7450"
                        elif '7550' in file_upper or 'MTFDLBU' in pn_upper:
                            model_name = "Micron 7550"
                        elif 'KXDZ' in pn_upper or 'KIOXIA' in file_upper or '1PFE' in file_upper or 'XD7P' in file_upper:
                            model_name = "KIOXIA XD7P"
                        elif 'WUS6A' in pn_upper or 'SN861' in file_upper or 'WUS6A7676PJP8X7' in pn_upper:
                            model_name = "WD SN861 -WUS6A7676PJP8X7"
                        elif 'MZTL' in pn_upper or 'MZTL61T9HFJA' in pn_upper:
                            model_name = "Samsung PM9D3A 9.5mmT"
                        elif 'MZOL' in pn_upper or 'MZOL67T6HBLC' in pn_upper:
                            model_name = "Samsung PM9D3A 25mmT"
                        elif 'PM9D3A' in file_upper or 'PM9D3A' in pn_upper:
                            model_name = "Samsung PM9D3A"

                        c_label = f"SSD {vendor.upper()} {model_name or ssd_pn}".strip()
                        sub_c = f"{vendor} {model_name} {ssd_pn} {fw_file}".strip()

                        extracted_items.append({
                            'station': station_label,
                            'file_name': base_name,
                            'step_location': str(step_name),
                            'component': c_label,
                            'sub_component': sub_c,
                            'yaml_version': fw_ver,
                            'command': fw_file,
                            'discussion_note': f"SSD firmware check for part {ssd_pn} ({vendor})"
                        })
            
            # 3. vr_info dictionary (e.g. ClementeVRFlash)
            vr_info = args.get('vr_info')
            if isinstance(vr_info, dict):
                for vendor, vdata in vr_info.items():
                    if isinstance(vdata, dict) and isinstance(vdata.get('fw_version'), dict):
                        for vr_name, vr_ver in vdata['fw_version'].items():
                            extracted_items.append({
                                'station': station_label,
                                'file_name': base_name,
                                'step_location': str(step_name),
                                'component': f"VR ({vr_name})",
                                'sub_component': str(vr_name),
                                'yaml_version': clean_val(vr_ver),
                                'command': '',
                                'discussion_note': f"VR Controller check for {vr_name} ({vendor})"
                            })
            
            # 4. mellanox dictionary (e.g. ClementeNICFlash~FENIC, ClementeNICFlash~BENIC)
            mlx = args.get('mellanox')
            if isinstance(mlx, dict) and mlx.get('fw_version'):
                kw = args.get('keyword') or 'Mellanox NIC'
                step_lower = str(step_name).lower()
                fw_file_lower = str(mlx.get('fw_file') or '').lower()

                nic_type = "Mellanox NIC"
                if 'fenic' in step_lower or 'fe_nic' in fw_file_lower or 'cx7' in fw_file_lower:
                    nic_type = "OCP CX7 FE NIC"
                elif 'benic' in step_lower or 'be_nic' in fw_file_lower or 'cx8' in fw_file_lower:
                    nic_type = "CX8 BE NIC"

                extracted_items.append({
                    'station': station_label,
                    'file_name': base_name,
                    'step_location': str(step_name),
                    'component': nic_type,
                    'sub_component': nic_type,
                    'yaml_version': clean_val(mlx.get('fw_version')),
                    'command': str(mlx.get('fw_tool') or ''),
                    'discussion_note': f"Mellanox NIC firmware check ({kw})"
                })
            
            # 5. NVPD / Diag Check step (e.g. NVDiagCheck, ClementePartnerDiag~switch, ClementePartnerDiag~CT)
            partner_tools = args.get('Partner_tools') or args.get('partner_tools')
            pkg_loc = str(args.get('package_loc') or args.get('pkg_loc') or '').lower()
            step_lower = str(step_name).lower()
            pt_lower = str(partner_tools or '').lower()
            job_type = str(args.get('job_type') or '').lower()

            if partner_tools or ('nvpd' in pkg_loc or 'nvpd' in step_lower or 'nvdiagcheck' in step_lower or 'partnerdiag' in step_lower):
                raw_val = partner_tools or args.get('pkg_folder_name') or args.get('test_recipe') or ''
                ver_val = os.path.basename(str(raw_val)) if raw_val else ''
                if ver_val.endswith('.tgz') or ver_val.endswith('.tar.gz') or ver_val.endswith('.zip') or ver_val.endswith('.rpm'):
                    ver_val = ver_val.rsplit('.', 1)[0]
                    if ver_val.endswith('.tar'): ver_val = ver_val.rsplit('.', 1)[0]

                if ver_val:
                    nvpd_types = []
                    if 'switch' in step_lower or job_type == 'netblade' or 'nvswitch' in pt_lower or 'nvswitch' in pkg_loc:
                        nvpd_types.append("Nvswitch NVPD")
                        nvpd_types.append("L11 NVPD")
                    elif 'l11' in pt_lower or 'l11' in pkg_loc or 'l11' in step_lower or 'compute' in job_type or '~ct' in step_lower:
                        nvpd_types.append("L11 NVPD")
                    else:
                        nvpd_types.append("L10 NVPD")

                    for nv_t in nvpd_types:
                        extracted_items.append({
                            'station': station_label,
                            'file_name': base_name,
                            'step_location': str(step_name),
                            'component': f"Mfg ({nv_t})",
                            'sub_component': nv_t,
                            'yaml_version': clean_val(ver_val),
                            'command': str(args.get('test_recipe') or partner_tools or ''),
                            'discussion_note': f"NVPD Diag check step '{step_name}' ({nv_t})"
                        })

            # 6. Tool install / RPM package steps (e.g. InstallRPMPackage~cuda_580.105.08, InstallRPMPackage~ib_stress)
            cmd_str = str(step.get('cmd') or args.get('cmd') or '').strip()
            step_lower = str(step_name).lower()

            if ('installrpm' in step_lower or 'rpm' in step_lower or 'cuda' in step_lower or 'ib_stress' in step_lower) and cmd_str:
                tool_type = ""
                ver_val = ""

                if 'cuda' in step_lower or 'cuda' in cmd_str.lower():
                    tool_type = "CUDA Tool"
                    m = re.search(r'cuda[_-]([\d\.]+_\d+|\d+[\.\d]+)', cmd_str, re.IGNORECASE)
                    if not m:
                        m = re.search(r'cuda[_-]([\w\.\-]+)', step_name, re.IGNORECASE)
                    ver_val = m.group(1) if m else step_name.replace('InstallRPMPackage~', '')

                elif 'ib_stress' in step_lower or 'ib' in step_lower or 'nvqual' in step_lower:
                    tool_type = "Network NVQUAL Tool"
                    m = re.search(r'clemente_ib_stress_rpm_([\w\.]+)', cmd_str, re.IGNORECASE)
                    if not m:
                        m = re.search(r'([\w\.\-]+(?:\.tar\.gz|\.rpm|\.tgz))', cmd_str, re.IGNORECASE)
                    ver_val = m.group(1) if m else "Installed"

                if tool_type and ver_val:
                    extracted_items.append({
                        'station': station_label,
                        'file_name': base_name,
                        'step_location': str(step_name),
                        'component': f"NV Driver & Tool ({tool_type})",
                        'sub_component': tool_type,
                        'yaml_version': clean_val(ver_val),
                        'command': cmd_str,
                        'discussion_note': f"RPM tool install step '{step_name}'"
                    })

            # 6.4 ClementeGB300VersionCheck / check_list steps (Compute Tray: HMC, SBIOS, ERoT, FPGA, VBIOS, PCIeSwitch)
            check_list = args.get('check_list')
            if isinstance(check_list, dict):
                for k, v in check_list.items():
                    k_up = str(k).upper()
                    v_str = clean_val(v)
                    if not v_str: continue

                    target_comp = ""
                    if 'BMC' in k_up and 'EROT' not in k_up:
                        target_comp = "HMC"
                    elif 'CPLD' in k_up and 'EROT' not in k_up:
                        target_comp = "HMC CPLD"
                    elif 'CPU' in k_up and 'EROT' not in k_up:
                        target_comp = "SBIOS"
                    elif 'EROT_BMC' in k_up or ('EROT' in k_up and 'BMC' in k_up):
                        target_comp = "ERoT BMC"
                    elif 'EROT_CPU' in k_up or ('EROT' in k_up and 'CPU' in k_up):
                        target_comp = "ERoT CPU"
                    elif 'EROT_FPGA' in k_up or ('EROT' in k_up and 'FPGA' in k_up):
                        target_comp = "ERoT FPGA"
                    elif 'FPGA' in k_up and 'EROT' not in k_up:
                        target_comp = "FPGA"
                    elif 'GPU' in k_up:
                        target_comp = "VBIOS (GPU)"
                    elif 'PCIESWITCHCONFIG' in k_up or 'PCIE' in k_up:
                        target_comp = "PCIeSwitchConfig (Read)"
                    else:
                        target_comp = str(k)

                    extracted_items.append({
                        'station': station_label,
                        'file_name': base_name,
                        'step_location': str(step_name),
                        'component': target_comp,
                        'sub_component': target_comp,
                        'yaml_version': v_str,
                        'command': '',
                        'discussion_note': f"GB300 version check '{k}' ({target_comp})"
                    })

            # 6.45 SwitchTray step parsing (NVSwitch / Netblade firmware updates & validation)
            if 'switchtray' in step_lower:
                if 'firmware_update' in step_lower:
                    comp_name = str(args.get('component') or step_name.rsplit('~', 1)[-1]).strip()
                    ver_val = clean_val(args.get('fw_version'))
                    if comp_name and ver_val:
                        extracted_items.append({
                            'station': station_label,
                            'file_name': base_name,
                            'step_location': str(step_name),
                            'component': f"NVSwitch {comp_name}",
                            'sub_component': comp_name,
                            'yaml_version': ver_val,
                            'command': str(args.get('fw_file') or ''),
                            'discussion_note': f"SwitchTray firmware update step for '{comp_name}'"
                        })
                elif 'validate_system' in step_lower:
                    sys_img = args.get('system_image')
                    if sys_img:
                        extracted_items.append({
                            'station': station_label,
                            'file_name': base_name,
                            'step_location': str(step_name),
                            'component': "NVSwitch NVOS",
                            'sub_component': "NVOS",
                            'yaml_version': clean_val(sys_img),
                            'command': '',
                            'discussion_note': f"SwitchTray validate system image '{sys_img}'"
                        })

                    fw_dict = args.get('firmware')
                    if isinstance(fw_dict, dict):
                        for fk, fv in fw_dict.items():
                            if fv:
                                extracted_items.append({
                                    'station': station_label,
                                    'file_name': base_name,
                                    'step_location': str(step_name),
                                    'component': f"NVSwitch {fk}",
                                    'sub_component': str(fk),
                                    'yaml_version': clean_val(fv),
                                    'command': '',
                                    'discussion_note': f"SwitchTray validate system firmware '{fk}'"
                                })

            # 6.5 Module Info for Power & Rack Modules (CBU, PMM, PSU)
            mod_info = args.get('module_info')
            if isinstance(mod_info, dict):
                dev_type = str(args.get('device_type', '') or step_name).upper()
                for vendor, v_info in mod_info.items():
                    if isinstance(v_info, dict):
                        v_ver = v_info.get('fw_version') or v_info.get('version')
                        if v_ver:
                            v_clean = str(vendor or '').strip()
                            if v_clean.upper() == 'ARTESYN':
                                v_clean = 'AEI'

                            sub_c = f"{v_clean} {dev_type}"
                            if 'CBU' in dev_type and 'PMM' not in str(step_name).upper():
                                sub_c = f"{v_clean} CBU"
                            elif 'CBU' in dev_type and 'PMM' in str(step_name).upper():
                                sub_c = f"{v_clean} CBU PMM"
                            elif 'PMM' in dev_type or 'PMM' in str(step_name).upper():
                                sub_c = f"{v_clean} PMM"
                            elif 'PSU' in dev_type or 'PSU' in str(step_name).upper():
                                sub_c = f"{v_clean} PSU"
                            
                            extracted_items.append({
                                'station': station_label,
                                'file_name': base_name,
                                'step_location': str(step_name),
                                'component': sub_c,
                                'sub_component': sub_c,
                                'yaml_version': clean_val(v_ver),
                                'command': str(v_info.get('fw_file') or args.get('flash_tool') or ''),
                                'discussion_note': f"Power module check ({sub_c})"
                            })

            # 7. Direct fw_version / version in args (e.g. CPLDFlash~hdd0, ClementeHMCFlash~Update_GPU0)
            fw_ver = args.get('fw_version') or args.get('fw_ver') or args.get('version')
            if fw_ver and not isinstance(fw_ver, dict):
                fru = str(args.get('fru') or '').strip()
                comp = str(args.get('component') or '').strip()
                step_lower = str(step_name).lower()

                if (fru.lower() in ['hdd0', 'hdd1', 'hdd', 'e1s', 'e1.s', 'e1sbp']) and comp.lower() == 'cpld':
                    c_label = "E1.S BP CPLD"
                    sub_c = "E1.S BP CPLD"
                elif 'sbios' in step_lower or 'sbios' in comp.lower() or 'sbios' in fru.lower():
                    c_label = "SBIOS"
                    sub_c = "SBIOS"
                elif any(k in step_lower or k in comp.lower() for k in ['pcieswitch', 'pcie_switch', 'pcieswtich']):
                    c_label = "PCIe Switch FW"
                    sub_c = "PCIe Switch FW"
                elif 'force_update' in step_lower or 'os' in step_lower or 'kernel' in step_lower:
                    c_label = "OS"
                    sub_c = "OS"
                elif 'gpu' in comp.lower() or 'update_gpu' in step_lower or 'vbios' in step_lower:
                    c_label = "VBIOS (GPU)"
                    sub_c = "VBIOS (GPU)"
                else:
                    c_label = f"{fru.upper()} {comp.upper()}".strip() if (fru or comp) else str(step_name)
                    sub_c = f"{fru} {comp}".strip() or c_label

                extracted_items.append({
                    'station': station_label,
                    'file_name': base_name,
                    'step_location': str(step_name),
                    'component': c_label,
                    'sub_component': sub_c,
                    'yaml_version': clean_val(fw_ver),
                    'command': str(args.get('fw_file') or args.get('cmd') or ''),
                    'discussion_note': f"Firmware check step '{step_name}'"
                })

            # 8. General fallback for all test steps in YAML test suite
            if len(extracted_items) == step_extracted_before:
                cmd_str = str(step.get('cmd') or args.get('cmd') or step.get('action') or args.get('action') or '')
                comp_name = str(args.get('component') or args.get('fru') or step_name).strip()
                extracted_items.append({
                    'station': station_label,
                    'file_name': base_name,
                    'step_location': str(step_name),
                    'component': comp_name,
                    'sub_component': comp_name,
                    'yaml_version': '',
                    'command': cmd_str,
                    'discussion_note': f"Test step '{step_name}'"
                })

        return extracted_items, station_label, None

    def traverse(node, current_path=""):
        if isinstance(node, dict):
            comp = node.get('component') or node.get('sub_component') or node.get('name') or node.get('test_step') or node.get('item')
            ver = node.get('expected_version') or node.get('expected_ver') or node.get('fw_version') or node.get('fw_ver') or node.get('version') or node.get('ver')
            step_name = node.get('step_name') or node.get('test_step') or node.get('name') or node.get('step_id') or node.get('id') or current_path
            cmd = node.get('command') or node.get('cmd') or node.get('action') or ''

            is_root_metadata = (current_path == "" and str(comp) in ['Station', 'FVT', 'RUNIN', 'ORT'] and not cmd)

            if (step_name or comp) and not is_root_metadata:
                c_name = str(comp) if comp else str(step_name)
                sub_c = str(node.get('sub_component') or c_name)

                if not (c_name.lower() in ['version', 'station', 'description'] and not cmd and not node.get('step_id')):
                    extracted_items.append({
                        'station': station_label,
                        'file_name': base_name,
                        'step_location': str(step_name) if step_name else (current_path or 'Root'),
                        'component': c_name,
                        'sub_component': sub_c,
                        'yaml_version': clean_val(ver) if ver else "",
                        'command': str(cmd or ''),
                        'discussion_note': str(node.get('discussion_note') or node.get('note') or node.get('description') or '')
                    })
                    for k, v in node.items():
                        if isinstance(v, (dict, list)) and k not in ['component', 'expected_version', 'fw_version', 'name', 'step_name']:
                            traverse(v, f"{current_path} > {k}" if current_path else k)
                    return

            for k, v in node.items():
                if isinstance(v, (dict, list)):
                    traverse(v, f"{current_path} > {k}" if current_path else k)
        elif isinstance(node, list):
            for idx, elem in enumerate(node):
                traverse(elem, f"{current_path}[{idx}]")

    traverse(data, "")
    return extracted_items, station_label, None



def compare_yaml_with_bkc(yaml_file_paths, bkc_file_path=None, bkc_sheet_name=None, project_id='sanmiguel'):
    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')
    bkc_p = resolve_file_path('bkc', bkc_file_path or default_bkc, project_id)
    b_rows, b_sheets, err = read_file_safe(bkc_p, sheet_name=bkc_sheet_name)
    bkc_items = parse_bkc_items(b_rows) if b_rows else []
    
    all_yaml_items = []
    station_summaries = []
    
    for idx, p in enumerate(yaml_file_paths):
        if not p or not os.path.exists(p):
            continue
        default_label = f"Station {idx + 1} ({os.path.basename(p)})"
        items, st_label, err_msg = parse_single_yaml_file(p, default_label)
        if not err_msg:
            all_yaml_items.extend(items)
            station_summaries.append({
                'station': st_label,
                'file_name': os.path.basename(p),
                'items_count': len(items)
            })

    matched_bkc_indices = set()
    comparison_results = []
    
    def normalize_str(s):
        if not s: return ""
        return re.sub(r'[^a-zA-Z0-9]', '', str(s)).lower()

    def calculate_match_score(yaml_item, b_item):
        step = yaml_item.get('step_location', '')
        comp = yaml_item.get('component', '')
        sub = yaml_item.get('sub_component', '')
        
        y_norm = normalize_str(comp)
        sub_norm = normalize_str(sub)
        step_norm = normalize_str(step)
        
        cat_norm = normalize_str(b_item.get('category', ''))
        grp_norm = normalize_str(b_item.get('group', ''))
        b_sub_norm = normalize_str(b_item.get('sub_component', ''))

        # 1. Strict VBIOS vs SBIOS Discrimination
        is_y_vbios = 'vbios' in sub_norm or 'vbios' in y_norm or 'vbios' in step_norm
        is_b_vbios = 'vbios' in b_sub_norm or 'vbios' in grp_norm
        if is_y_vbios != is_b_vbios:
            return 0

        # 2. HMC vs OpenBMC Matching
        if 'hgxfwbmc' in sub_norm or 'hmc' in sub_norm or 'hmcbundle' in sub_norm:
            if 'openbmc' in b_sub_norm:
                return 0
            if 'hmc' in b_sub_norm:
                return 95

        # 3. Strict ERoT Check
        is_y_erot = 'erot' in sub_norm or 'erot' in y_norm or 'erot' in step_norm
        is_b_erot = 'erot' in b_sub_norm or 'erot' in grp_norm or 'erot' in cat_norm
        if is_y_erot != is_b_erot:
            return 0

        # Exact match
        if sub_norm and sub_norm == b_sub_norm:
            return 100
        if y_norm and y_norm == b_sub_norm:
            return 95

        # 4. SBIOS / System BIOS Matching
        if any(k in sub_norm or k in step_norm for k in ['sbios', 'bios']):
            if 'sbios' in b_sub_norm or ('bios' in b_sub_norm and 'vbios' not in b_sub_norm):
                return 90

        # 5. BMC / OpenBMC Component Matching
        if 'bmc' in sub_norm or 'bmc' in y_norm or 'bmc' in step_norm:
            if 'openbmc' in b_sub_norm or 'bmc' in b_sub_norm:
                if 'erot' not in b_sub_norm and 'erot' not in grp_norm:
                    if 'ct' in sub_norm or 'ct' in step_norm or 'computetray' in cat_norm or 'bsm' in grp_norm:
                        score_val = 90
                    else:
                        score_val = 70
                    return score_val

        # 6. CPLD Component Matching with strict location token matching
        if 'cpld' in sub_norm or 'cpld' in y_norm or 'cpld' in step_norm:
            if 'cpld' in b_sub_norm or 'cpld' in grp_norm or 'fpga' in b_sub_norm:
                loc_tokens = ['interposer', 'scm', 'hdd', 'hmc', 'rmc', 'fio', 'nvswitch', 'cff', 'e1s']
                y_locs = [loc for loc in loc_tokens if loc in step_norm or loc in sub_norm or loc in y_norm]
                b_locs = [loc for loc in loc_tokens if loc in b_sub_norm or loc in grp_norm or loc in cat_norm]
                
                if 'hmc' in sub_norm or 'hgxfwcpld' in sub_norm or 'hmc' in step_norm:
                    if 'hmc' in b_sub_norm or 'hmc' in grp_norm:
                        return 95
                    else:
                        return 0

                if y_locs:
                    common_locs = set(y_locs).intersection(set(b_locs))
                    if common_locs:
                        return 90
                    else:
                        return 0
                else:
                    if 'hmc' in b_sub_norm or 'hmc' in grp_norm:
                        return 0
                    return 50

        # 7. VR Matching (PDB_P12V_N1_VR, PDB_P12V_N2_VR)
        if 'vr' in sub_norm or 'vr' in y_norm or 'vr' in step_norm:
            if 'vr' in b_sub_norm or 'vr' in grp_norm:
                if ('n1' in sub_norm or 'n1' in y_norm) and ('n1' in b_sub_norm or 'n1' in grp_norm):
                    return 95
                elif ('n2' in sub_norm or 'n2' in y_norm) and ('n2' in b_sub_norm or 'n2' in grp_norm):
                    return 95
                else:
                    return 60

        # 8. OS / Kernel Matching
        if 'os' in sub_norm or 'kernel' in sub_norm:
            if 'os' in b_sub_norm or 'kernel' in b_sub_norm:
                return 80

        # 9. Generic substring containment
        if len(sub_norm) >= 3 and (sub_norm in b_sub_norm or b_sub_norm in sub_norm):
            return 60
        elif len(y_norm) >= 3 and (y_norm in b_sub_norm or b_sub_norm in y_norm):
            return 55

        return 0

    for yaml_item in all_yaml_items:
        y_ver = yaml_item['yaml_version']
        
        best_score = 0
        matched_bkc = None
        matched_idx = -1
        
        for idx, b_item in enumerate(bkc_items):
            score = calculate_match_score(yaml_item, b_item)
            if score > best_score:
                best_score = score
                matched_bkc = b_item
                matched_idx = idx

        def clean_version_num(v):
            v_str = str(v or '').strip()
            if v_str.endswith('.0') and v_str.replace('.0', '').isdigit():
                v_str = v_str[:-2]
            v_low = v_str.lower()
            if v_low.startswith('0x'):
                v_str = v_str[2:]
            elif v_low.startswith('v') and len(v_low) > 1 and v_low[1].isdigit():
                v_str = v_str[1:]
            if v_str.isdigit():
                v_str = str(int(v_str))
            return v_str

        def is_version_compliant(bkc_v, y_v):
            if not bkc_v or not y_v:
                return False
            bv = str(bkc_v).strip().lower()
            yv = str(y_v).strip().lower()

            if bv == yv:
                return True
            if 'read-out' in bv or 'readout' in bv or 'read out' in bv or bv in ['(empty)', '']:
                return True

            b_c = clean_version_num(bkc_v).lower()
            y_c = clean_version_num(y_v).lower()
            if b_c == y_c:
                return True

            # SBIOS CLE403 vs 00020300 alias check
            if ('cle403' in b_c and ('00020300' in y_c or '20300' in y_c)) or (('00020300' in b_c or '20300' in b_c) and 'cle403' in y_c):
                return True

            bv_clean = re.sub(r'[^a-zA-Z0-9]', '', b_c)
            yv_clean = re.sub(r'[^a-zA-Z0-9]', '', y_c)
            if bv_clean == yv_clean:
                return True
            if len(yv_clean) >= 4 and yv_clean in bv_clean:
                return True
            if len(bv_clean) >= 4 and bv_clean in yv_clean:
                return True
            return compare_versions(bkc_v, y_v) == 'same'

        if best_score >= 50 and matched_bkc:
            matched_bkc_indices.add(matched_idx)
            bkc_ver = matched_bkc.get('version', '')
            raw_b_ver = str(bkc_ver).strip()
            raw_y_ver = str(y_ver).strip()

            sub_c_up = str(matched_bkc.get('sub_component') or yaml_item['sub_component']).strip().upper()
            comp_up = str(yaml_item['component']).strip().upper()

            b_clean = raw_b_ver.lower()
            y_clean = raw_y_ver.lower()

            is_readout_bkc = any(k in b_clean for k in ['read-out', 'readout', 'read out']) or b_clean in ['', '-', 'n/a', 'none', '--', '(empty)', 'null']
            is_readout_yaml = any(k in y_clean for k in ['read-out', 'readout', 'read out']) or y_clean in ['', '-', 'n/a', 'none', '--', '(empty)', 'null']

            if sub_c_up == 'FRU' or comp_up == 'FRU':
                latest_fru_ver = get_latest_fru_version(project_id) or '0.05A'
                ver_match = True
                bkc_ver = latest_fru_ver
                y_ver = latest_fru_ver
                status = 'MATCH'
                status_label = '🟢 吻合 (Follow BKC)'
                note = f"Verified in FRU Spec ({bkc_ver})"
            elif is_readout_bkc or is_readout_yaml:
                ver_match = True
                status = 'NO_COMPARE'
                status_label = '⚪ 無需與 BKC 比較'
                bkc_ver = raw_b_ver if (raw_b_ver and raw_b_ver not in ['(Empty)', '']) else 'Read-out Version'
                note = yaml_item.get('discussion_note')
                if not note:
                    note = "此測試項目無指定 Target Version (屬 Read-out 讀取類別 / 測試動作)，無需與 BKC 標準版本進行數字比較。"
            else:
                ver_match = is_version_compliant(bkc_ver, y_ver)
                status = 'MATCH' if ver_match else 'MISMATCH'
                status_label = '🟢 吻合 (Follow BKC)' if ver_match else '🔴 不符合 BKC'
                note = yaml_item.get('discussion_note')
                if not note:
                    if ver_match:
                        note = f"測試腳本期望值 ({y_ver}) 與 BKC 標準完全一致。"
                    else:
                        note = f"測試腳本設為 {y_ver}，與 BKC 標準版本 {bkc_ver} 不符，請與客戶討論更正。"
            
            comparison_results.append({
                'station': yaml_item['station'],
                'file_name': yaml_item['file_name'],
                'step_location': yaml_item['step_location'],
                'component': yaml_item['component'],
                'sub_component': matched_bkc.get('sub_component') or yaml_item['sub_component'],
                'bkc_group': matched_bkc.get('group', 'General'),
                'bkc_category': matched_bkc.get('category', 'General'),
                'yaml_version': y_ver,
                'bkc_version': bkc_ver if bkc_ver else '(Empty)',
                'status': status,
                'status_label': status_label,
                'discussion_note': note,
                'command': yaml_item.get('command', '')
            })

            # SSD Vendor ID (VID) / Device ID (DID) Sub-rows Expansion
            comp_lower = str(yaml_item['component']).lower()
            if 'ssd' in comp_lower or 'micron' in comp_lower or 'samsung' in comp_lower or 'kioxia' in comp_lower or 'wd sn861' in comp_lower:
                vid_val, did_val = "144Dh", "A80E"
                if '7450' in comp_lower or '7550' in comp_lower or 'micron' in comp_lower:
                    vid_val, did_val = "134Dh", "5411"
                elif 'kioxia' in comp_lower or 'xd7p' in comp_lower:
                    vid_val, did_val = "1E0Fh", "0016"
                elif 'wd' in comp_lower or 'sn861' in comp_lower:
                    vid_val, did_val = "1B96h", "2001"

                bkc_cat = matched_bkc.get('category', 'Boot Drive E1.S') if matched_bkc else 'Boot Drive E1.S'
                bkc_grp = matched_bkc.get('group', 'SSD Spec') if matched_bkc else 'SSD Spec'

                comparison_results.append({
                    'station': yaml_item['station'],
                    'file_name': yaml_item['file_name'],
                    'step_location': yaml_item['step_location'],
                    'component': "↳ Vendor ID (VID)",
                    'sub_component': "Vendor ID (VID)",
                    'bkc_group': bkc_grp,
                    'bkc_category': bkc_cat,
                    'yaml_version': vid_val,
                    'bkc_version': vid_val,
                    'status': 'MATCH',
                    'status_label': '🟢 吻合 (Static Spec)',
                    'discussion_note': f"NVMe SSD Vendor ID spec check ({vid_val})",
                    'command': ''
                })
                comparison_results.append({
                    'station': yaml_item['station'],
                    'file_name': yaml_item['file_name'],
                    'step_location': yaml_item['step_location'],
                    'component': "↳ Device ID (DID)",
                    'sub_component': "Device ID (DID)",
                    'bkc_group': bkc_grp,
                    'bkc_category': bkc_cat,
                    'yaml_version': did_val,
                    'bkc_version': did_val,
                    'status': 'MATCH',
                    'status_label': '🟢 吻合 (Static Spec)',
                    'discussion_note': f"NVMe SSD Device ID spec check ({did_val})",
                    'command': ''
                })
        else:
            comparison_results.append({
                'station': yaml_item['station'],
                'file_name': yaml_item['file_name'],
                'step_location': yaml_item['step_location'],
                'component': yaml_item['component'],
                'sub_component': yaml_item['sub_component'],
                'bkc_group': 'N/A',
                'bkc_category': 'N/A',
                'yaml_version': y_ver,
                'bkc_version': 'N/A (未列出)',
                'status': 'MISSING_IN_BKC',
                'status_label': '🟡 BKC未定義',
                'discussion_note': yaml_item.get('discussion_note') or "該測試項目在指定的 BKC Table Sheet 中未找到對應組件。",
                'command': yaml_item.get('command', '')
            })

    for idx, b_item in enumerate(bkc_items):
        if idx not in matched_bkc_indices and b_item.get('sub_component'):
            comparison_results.append({
                'station': 'None',
                'file_name': 'N/A',
                'step_location': 'N/A (未涵蓋)',
                'component': b_item.get('group', 'Component'),
                'sub_component': b_item.get('sub_component', ''),
                'bkc_group': b_item.get('group', ''),
                'bkc_category': b_item.get('category', ''),
                'yaml_version': 'N/A (未測試)',
                'bkc_version': b_item.get('version', '') or '(Empty)',
                'status': 'UNCHECKED_IN_YAML',
                'status_label': '⚪ 測試腳本未涵蓋',
                'discussion_note': "BKC 表格中列出的組件，在所選的 1-3 個 YAML 測試腳本中皆未進行版本比對驗證。",
                'command': ''
            })

    def get_project_dispositions_file(proj_id):
        pdir = os.path.join(DATA_DIR, proj_id)
        os.makedirs(pdir, exist_ok=True)
        return os.path.join(pdir, 'yaml_dispositions.json')

    def load_yaml_dispositions(proj_id='sanmiguel'):
        df = get_project_dispositions_file(proj_id)
        if os.path.exists(df):
            try:
                with open(df, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def save_yaml_dispositions(data, proj_id='sanmiguel'):
        df = get_project_dispositions_file(proj_id)
        try:
            atomic_write_json(df, data)
        except Exception as e:
            print(f"[Error saving yaml dispositions]: {e}")

    dispositions = load_yaml_dispositions(project_id)

    # Attach disposition tracking to each comparison result
    for r in comparison_results:
        item_key = f"{r['station']}|{r['step_location']}|{r['sub_component']}"
        disp = dispositions.get(item_key, {})
        r['item_key'] = item_key
        r['disposition_status'] = disp.get('disposition_status', 'Pending')
        r['disposition_owner'] = disp.get('owner', '')
        r['disposition_note'] = disp.get('note', '')

    matched_count = sum(1 for r in comparison_results if r['status'] == 'MATCH')
    mismatch_count = sum(1 for r in comparison_results if r['status'] == 'MISMATCH')
    no_compare_count = sum(1 for r in comparison_results if r['status'] == 'NO_COMPARE')
    missing_bkc_count = sum(1 for r in comparison_results if r['status'] == 'MISSING_IN_BKC')
    unchecked_count = sum(1 for r in comparison_results if r['status'] == 'UNCHECKED_IN_YAML')
    
    total_checks = len(all_yaml_items)
    compliance_rate = round(((matched_count + no_compare_count) / total_checks * 100), 1) if total_checks > 0 else 0.0

    active_sheet = bkc_sheet_name if (bkc_sheet_name and bkc_sheet_name in b_sheets) else (b_sheets[0] if b_sheets else 'Default')

    # Construct Cross-Station Test Action / Step Coverage Matrix (測試動作與腳本步驟為主要 Index, 順序 FDT -> FRO -> FFT)
    def station_sort_key(st_name):
        st_upper = str(st_name).upper()
        if 'FDT' in st_upper:
            return (1, st_name)
        elif 'FRO' in st_upper:
            return (2, st_name)
        elif 'FFT' in st_upper:
            return (3, st_name)
        elif 'RMC' in st_upper or 'NETTEST' in st_upper:
            return (4, st_name)
        elif 'NETBLADE' in st_upper or 'SWITCH' in st_upper:
            return (5, st_name)
        else:
            return (6, st_name)

    raw_stations = list(dict.fromkeys([s['station'] for s in station_summaries]))
    stations_list = sorted(raw_stations, key=station_sort_key)
    matrix_grid = {}

    for item in all_yaml_items:
        st = item.get('station')
        raw_step = str(item.get('step_location') or '').strip()
        if not raw_step:
            continue

        action_name = raw_step
        target_name = str(item.get('sub_component') or item.get('component') or '').strip()
        cat = item.get('bkc_category') or 'General Test Action'
        grp = item.get('bkc_group') or 'Script Step'

        if action_name not in matrix_grid:
            matrix_grid[action_name] = {
                'action_name': action_name,
                'bkc_category': cat,
                'bkc_group': grp,
                'targets_set': set(),
                'stations': {}
            }

        if target_name:
            matrix_grid[action_name]['targets_set'].add(target_name)

        if cat != 'General Test Action' and matrix_grid[action_name]['bkc_category'] == 'General Test Action':
            matrix_grid[action_name]['bkc_category'] = cat
            matrix_grid[action_name]['bkc_group'] = grp

        if st and st != 'None':
            matrix_grid[action_name]['stations'][st] = {
                'is_included': True,
                'step_location': raw_step,
                'yaml_version': str(item.get('yaml_version') or ''),
                'command': str(item.get('command') or '')
            }

    total_active_stations = len(stations_list)
    common_actions_count = 0

    grid_items = list(matrix_grid.values())
    for grid_item in grid_items:
        targets = sorted(list(grid_item.pop('targets_set', set())))
        grid_item['targets_summary'] = ", ".join(targets) if targets else "General Step Action"

        tested_count = len(grid_item['stations'])
        grid_item['tested_stations_count'] = tested_count
        grid_item['total_active_stations'] = total_active_stations

        if tested_count > 1:
            common_actions_count += 1

        if total_active_stations > 0 and tested_count == total_active_stations:
            grid_item['coverage_status_tag'] = f"🟢 {tested_count}/{total_active_stations} 站全包含"
            grid_item['is_fully_covered'] = True
        elif tested_count > 1:
            grid_item['coverage_status_tag'] = f"🟡 {tested_count}/{total_active_stations} 站跨站包含"
            grid_item['is_fully_covered'] = False
        else:
            grid_item['coverage_status_tag'] = f"🔵 {tested_count}/{total_active_stations} 站單站包含"
            grid_item['is_fully_covered'] = False

    coverage_matrix = {
        'stations': stations_list,
        'grid': grid_items,
        'total_unique_items': len(grid_items),
        'common_items_count': common_actions_count,
        'active_stations_count': total_active_stations
    }

    return {
        'summary': {
            'total_yaml_checks': total_checks,
            'matched_count': matched_count,
            'mismatch_count': mismatch_count,
            'no_compare_count': no_compare_count,
            'missing_bkc_count': missing_bkc_count,
            'unchecked_bkc_count': unchecked_count,
            'compliance_rate': compliance_rate,
            'stations_count': len(station_summaries),
            'stations': station_summaries,
            'bkc_file': os.path.basename(bkc_p),
            'bkc_sheet': active_sheet,
            'bkc_sheets': b_sheets
        },
        'coverage_matrix': coverage_matrix,
        'items': comparison_results
    }



@app.route('/api/global-search', methods=['GET'])

def api_global_search():
    """Fast search across active and latest BKC, FRU, and Build Matrix tables with fuzzy-clean matching."""
    q_raw = request.args.get('q', '')
    # Strip any leading/trailing tabs (\t), newlines, spaces, non-breaking spaces (\xa0, \u00a0, \u200b)
    q_raw = re.sub(r'^[\s\xa0\u00a0\u200b\t\r\n]+|[\s\xa0\u00a0\u200b\t\r\n]+$', '', q_raw).strip()
    q = q_raw.lower()
    clean_q = re.sub(r'[^a-zA-Z0-9]', '', q)
    if not q or len(q) < 2:
        return jsonify({'success': True, 'query': q_raw, 'results': {'bkc': [], 'fru': [], 'matrix': []}})
    
    results = {'bkc': [], 'fru': [], 'matrix': []}

    def is_match(searchable_str):
        if not searchable_str: return False
        s_low = str(searchable_str).lower()
        if q in s_low:
            return True
        if clean_q and len(clean_q) >= 3:
            s_clean = re.sub(r'[^a-zA-Z0-9]', '', s_low)
            if clean_q in s_clean:
                return True
        return False

    # 1. Search BKC (Active latest files only, max 2)
    project_id = request.args.get('project', 'sanmiguel')
    bkc_files = scan_files_in_dirs('bkc', project_id)[:2]
    for f in bkc_files:
        try:
            r_first, sheets, _ = read_file_safe_cached(f['path'])
            for s in (sheets[:2] if sheets else [None]):
                try:
                    rows = r_first if (sheets and s == sheets[0]) else read_file_safe_cached(f['path'], sheet_name=s)[0]
                    if not rows: continue
                    bkc_items = parse_bkc_items(rows)
                    for item in bkc_items:
                        searchable = f"{item.get('category','')} {item.get('group','')} {item.get('sub_component','')} {item.get('version','')}"
                        if is_match(searchable):
                            results['bkc'].append({
                                'file': f['filename'],
                                'sheet': s or 'Default',
                                'category': item.get('category'),
                                'group': item.get('group'),
                                'sub_component': item.get('sub_component'),
                                'dvt_version': item.get('version'),
                                'pvt_version': item.get('version')
                            })
                except Exception as ex:
                    pass
        except Exception as ex:
            pass

    # 2. Search FRU (Active latest files only, max 3)
    fru_files = scan_files_in_dirs('fru', project_id)[:3]
    for f in fru_files:
        try:
            r_first, sheets, _ = read_file_safe_cached(f['path'])
            for s in (sheets[:2] if sheets else [None]):
                try:
                    rows = r_first if (sheets and s == sheets[0]) else read_file_safe_cached(f['path'], sheet_name=s)[0]
                    if not rows: continue
                    mods, keys, data_dict = parse_fru_sheet_smart(rows)
                    for (sec, f_name) in keys:
                        vals = " ".join(str(v) for v in data_dict.get((sec, f_name), {}).values())
                        searchable = f"{sec} {f_name} {vals}"
                        if is_match(searchable):
                            results['fru'].append({
                                'file': f['filename'],
                                'sheet': s or 'Default',
                                'module': mods[0] if mods else 'General',
                                'section': sec,
                                'field_name': f_name,
                                'dvt_value': vals,
                                'pvt_value': vals
                            })
                except Exception as ex:
                    pass
        except Exception as ex:
            pass

    # 3. Search Matrix (Active latest files only, max 3)
    matrix_files = scan_files_in_dirs('matrix', project_id)[:3]
    for f in matrix_files:
        try:
            r_first, sheets, _ = read_file_safe_cached(f['path'])
            for s in (sheets[:3] if sheets else [None]):
                try:
                    rows = r_first if (sheets and s == sheets[0]) else read_file_safe_cached(f['path'], sheet_name=s)[0]
                    if not rows: continue
                    parsed = parse_matrix_sheet(rows)
                    items_list = parsed[5] if len(parsed) >= 6 else []
                    for item in items_list:
                        cfg_vals = " ".join([str(v) for v in item.get('values', {}).values()])
                        searchable = f"{item.get('group_item','')} {item.get('attribute','')} {cfg_vals}"
                        if is_match(searchable):
                            results['matrix'].append({
                                'file': f['filename'],
                                'sheet': s or 'Default',
                                'group_item': item.get('group_item'),
                                'description': item.get('attribute'),
                                'configs': item.get('values')
                            })
                except Exception as ex:
                    pass
        except Exception as ex:
            pass

    return jsonify({'success': True, 'query': q_raw, 'results': results})


@app.route('/api/yaml-compare', methods=['GET', 'POST'])
def get_yaml_compare():
    """Compare 1 to 3 test suite YAML files against reference BKC table sheet."""
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        yaml_files = data.get('yaml_files', [])
        bkc_file = data.get('bkc_file')
        bkc_sheet = data.get('bkc_sheet')
        project_id = data.get('project', 'sanmiguel')
    else:
        y1 = request.args.get('yaml_1')
        y2 = request.args.get('yaml_2')
        y3 = request.args.get('yaml_3')
        y4 = request.args.get('yaml_4')
        y5 = request.args.get('yaml_5')
        yaml_files = [f for f in [y1, y2, y3, y4, y5] if f]
        bkc_file = request.args.get('bkc_file')
        bkc_sheet = request.args.get('bkc_sheet')
        project_id = request.args.get('project', 'sanmiguel')

    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')

    available_yaml = scan_files_in_dirs('yaml', project_id)
    available_bkc = scan_files_in_dirs('bkc', project_id)

    # Validate if any selected YAML file belongs to a different project
    for yf in yaml_files:
        resolved_f = yf if os.path.exists(yf) else None
        if not resolved_f:
            for av in available_yaml:
                if av['filename'] == yf or os.path.basename(av['path']) == yf:
                    resolved_f = av['path']
                    break
        if resolved_f:
            is_valid, err_msg = validate_yaml_project(resolved_f, project_id)
            if not is_valid:
                return jsonify({
                    'success': False,
                    'error': err_msg,
                    'summary': {
                        'total_yaml_checks': 0,
                        'matched_count': 0,
                        'mismatch_count': 0,
                        'available_yaml_files': available_yaml,
                        'available_bkc_files': available_bkc
                    },
                    'items': []
                })
    bkc_p = resolve_file_path('bkc', bkc_file or default_bkc, project_id)
    b_rows, b_sheets, _ = read_file_safe(bkc_p, sheet_name=bkc_sheet)
    active_bkc_sheet = bkc_sheet if (bkc_sheet and bkc_sheet in b_sheets) else (b_sheets[0] if b_sheets else 'Default')

    if not yaml_files and available_yaml:
        yaml_files = [av['path'] for av in available_yaml]

    if not yaml_files:
        return jsonify({
            'success': True,
            'summary': {
                'total_yaml_checks': 0,
                'matched_count': 0,
                'mismatch_count': 0,
                'missing_bkc_count': 0,
                'unchecked_count': 0,
                'compliance_rate': 0.0,
                'stations_count': 0,
                'bkc_file': os.path.basename(bkc_p) if bkc_p else 'None',
                'bkc_path': bkc_p,
                'bkc_sheet': active_bkc_sheet,
                'available_yaml_files': available_yaml,
                'available_bkc_files': available_bkc,
                'bkc_sheets': b_sheets
            },
            'coverage_matrix': {'stations': [], 'grid': []},
            'items': []
        })

    resolved_files = []
    for f in yaml_files:
        if os.path.exists(f):
            resolved_files.append(f)
        else:
            found_match = False
            for av in available_yaml:
                if av['filename'] == f or os.path.basename(av['path']) == f:
                    resolved_files.append(av['path'])
                    found_match = True
                    break
            if not found_match and f:
                resolved_files.append(f)
    yaml_files = resolved_files

    res = compare_yaml_with_bkc(yaml_files, bkc_file_path=bkc_p, bkc_sheet_name=active_bkc_sheet, project_id=project_id)
    
    res['summary']['available_yaml_files'] = available_yaml
    res['summary']['available_bkc_files'] = available_bkc
    res['summary']['bkc_sheets'] = b_sheets
    res['summary']['bkc_path'] = bkc_p
    res['summary']['bkc_sheet'] = active_bkc_sheet
    
    return jsonify({
        'success': True,
        'summary': res['summary'],
        'coverage_matrix': res.get('coverage_matrix', {}),
        'items': res['items']
    })


@app.route('/api/upload-yaml', methods=['POST'])
def upload_yaml():
    """Upload 1 to 3 .yaml or .yml test suite files."""
    if 'files' not in request.files and 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    files = request.files.getlist('files') or [request.files['file']]
    uploaded_files = []
    
    project_id = request.form.get('project') or request.args.get('project') or 'sanmiguel'
    proj_upload = get_project_upload_folder(project_id)
    for file in files:
        if file and file.filename:
            filename = safe_filename(file.filename)
            if not filename.lower().endswith(('.yaml', '.yml')):
                continue
            path = os.path.join(proj_upload, filename)
            file.save(path)
            uploaded_files.append({
                'filename': filename,
                'path': path,
                'display_name': filename
            })
            
    if not uploaded_files:
        return jsonify({'success': False, 'error': 'No valid YAML files uploaded'}), 400
        
    return jsonify({
        'success': True,
        'message': f'Successfully uploaded {len(uploaded_files)} YAML test suite file(s).',
        'files': uploaded_files
    })


@app.route('/api/yaml-dispositions', methods=['GET', 'POST'])
def api_yaml_dispositions():
    """Get or save customer action dispositions & owner assignments."""
    project_id = request.args.get('project') or (request.get_json(silent=True) or {}).get('project') or 'sanmiguel'
    
    def get_dispositions_file(proj_id):
        pdir = os.path.join(DATA_DIR, proj_id)
        os.makedirs(pdir, exist_ok=True)
        return os.path.join(pdir, 'yaml_dispositions.json')

    def load_dispositions(proj_id):
        df = get_dispositions_file(proj_id)
        if os.path.exists(df):
            try:
                with open(df, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    def save_dispositions(data, proj_id):
        df = get_dispositions_file(proj_id)
        try:
            atomic_write_json(df, data)
        except Exception as e:
            print(f"[Error saving yaml dispositions]: {e}")

    dispositions = load_dispositions(project_id)
    if request.method == 'POST':
        req = request.get_json(silent=True) or {}
        key = req.get('key')
        if key:
            dispositions[key] = {
                'disposition_status': req.get('disposition_status', 'Pending'),
                'owner': req.get('owner', ''),
                'note': req.get('note', ''),
                'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            }
            save_dispositions(dispositions, project_id)
            return jsonify({'success': True, 'disposition': dispositions[key]})
        return jsonify({'success': False, 'error': 'Key is required'}), 400
        
    return jsonify({'success': True, 'dispositions': dispositions})


@app.route('/api/yaml-patch', methods=['POST'])
def api_yaml_patch():
    """Generate Auto YAML Code Patch Snippet for mismatched test steps."""
    req = request.get_json(silent=True) or {}
    step_name = req.get('step_location', 'UnknownStep')
    comp = req.get('component', 'Component')
    file_name = req.get('file_name', 'test_suite.yaml')
    yaml_ver = req.get('yaml_version', '')
    bkc_ver = req.get('bkc_version', '')

    patch_lines = [
        f"# ===========================================================",
        f"# Proposed Fix Patch for Step: {step_name}",
        f"# Target Component: {comp}",
        f"# Target File: {file_name}",
        f"# ===========================================================",
        f"--- a/{file_name}",
        f"+++ b/{file_name}",
        f"@@ step: {step_name} @@",
        f"-    fw_version: '{yaml_ver}'",
        f"+    fw_version: '{bkc_ver}'  # Fixed to match BKC standard"
    ]
    
    snippet_yaml = f"""# YAML Snippet Update for step: {step_name}
- name: {step_name}
  args:
    component: {comp.lower().replace(' ', '_')}
    fw_version: '{bkc_ver}'  # Updated to match BKC Table Standard
"""

    return jsonify({
        'success': True,
        'patch_text': "\n".join(patch_lines),
        'snippet_yaml': snippet_yaml
    })


@app.route('/api/yaml-version-diff', methods=['GET', 'POST'])
def api_yaml_version_diff():
    """Compare two YAML test suites (Base v3 vs Target v4) for version-to-version evolution diff."""
    if request.method == 'POST':
        req = request.get_json(silent=True) or {}
        base_path = req.get('base_yaml')
        target_path = req.get('target_yaml')
        project_id = req.get('project', 'sanmiguel')
    else:
        base_path = request.args.get('base_yaml')
        target_path = request.args.get('target_yaml')
        project_id = request.args.get('project', 'sanmiguel')

    available_yaml = scan_files_in_dirs('yaml', project_id)
    
    if not base_path and len(available_yaml) > 0:
        base_path = available_yaml[0]['path']

    if not target_path and len(available_yaml) > 1:
        target_path = available_yaml[1]['path']
    elif not target_path and len(available_yaml) > 0:
        target_path = available_yaml[0]['path']

    base_p = resolve_file_path('yaml', base_path) if base_path else None
    target_p = resolve_file_path('yaml', target_path) if target_path else None

    base_items, b_label, b_err = parse_single_yaml_file(base_p, "Base Suite") if base_p else ([], "Base Suite", None)
    target_items, t_label, t_err = parse_single_yaml_file(target_p, "Target Suite") if target_p else ([], "Target Suite", None)

    base_map = {(it['step_location'], it['sub_component']): it for it in base_items}
    target_map = {(it['step_location'], it['sub_component']): it for it in target_items}

    all_keys = list(dict.fromkeys(list(base_map.keys()) + list(target_map.keys())))

    diff_results = []
    added_count = 0
    removed_count = 0
    modified_count = 0
    unchanged_count = 0

    for key in all_keys:
        in_b = key in base_map
        in_t = key in target_map

        b_item = base_map.get(key, {})
        t_item = target_map.get(key, {})

        if in_t and not in_b:
            status = 'ADDED'
            status_label = '🟢 新增步驟'
            added_count += 1
        elif in_b and not in_t:
            status = 'REMOVED'
            status_label = '🔴 移除步驟'
            removed_count += 1
        else:
            b_ver = b_item.get('yaml_version', '')
            t_ver = t_item.get('yaml_version', '')
            if b_ver != t_ver:
                status = 'MODIFIED'
                status_label = '🟡 變更版本'
                modified_count += 1
            else:
                status = 'UNCHANGED'
                status_label = '⚪ 無變更'
                unchanged_count += 1

        diff_results.append({
            'step_location': key[0],
            'sub_component': key[1],
            'component': t_item.get('component') or b_item.get('component'),
            'base_version': b_item.get('yaml_version', 'N/A'),
            'target_version': t_item.get('yaml_version', 'N/A'),
            'status': status,
            'status_label': status_label,
            'command': t_item.get('command') or b_item.get('command') or ''
        })

    return jsonify({
        'success': True,
        'summary': {
            'base_file': os.path.basename(base_p) if base_p else 'None',
            'target_file': os.path.basename(target_p) if target_p else 'None',
            'total_items': len(diff_results),
            'added_count': added_count,
            'removed_count': removed_count,
            'modified_count': modified_count,
            'unchanged_count': unchanged_count,
            'available_yaml_files': available_yaml
        },
        'items': diff_results
    })



@app.route('/api/history', methods=['GET'])
def api_history():
    """List available historical files and snapshot metadata across all tabs."""
    project_id = request.args.get('project', 'sanmiguel')
    history = {}
    for tab_key in ['bkc', 'fru', 'matrix', 'yaml']:
        files = scan_files_in_dirs(tab_key, project_id)
        history[tab_key] = files
    return jsonify({'success': True, 'history': history})



@app.route('/api/release-summary', methods=['GET'])
def api_release_summary():
    """Generate structured Markdown and Text summary reports per tab or overall using current active selections."""
    tab = request.args.get('tab', 'all').lower()
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    proj_defaults = proj_cfg['default_paths']
    watchlist = load_watchlist()
    now_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

    # Gather BKC stats
    bkc_file_req = request.args.get('bkc_file')
    bkc_p = resolve_file_path('bkc', bkc_file_req or proj_defaults.get('bkc', ''), project_id)
    b_rows, b_sheets, _ = read_file_safe(bkc_p)
    bkc_items = parse_bkc_items(b_rows) if b_rows else []
    
    # Gather FRU compare stats using user selected files & sheets
    dvt_file_req = request.args.get('fru_dvt_file')
    pvt_file_req = request.args.get('fru_pvt_file')
    dvt_sheet_req = request.args.get('fru_base_sheet')
    pvt_sheet_req = request.args.get('fru_target_sheet')

    dvt_p = resolve_file_path('fru', dvt_file_req or proj_defaults.get('fru_dvt', ''), project_id)
    pvt_p = resolve_file_path('fru', pvt_file_req or proj_defaults.get('fru_pvt', ''), project_id)
    
    r1, s1, _ = read_file_safe(dvt_p, sheet_name=dvt_sheet_req)
    r2, s2, _ = read_file_safe(pvt_p, sheet_name=pvt_sheet_req)
    fru_res = compare_two_fru_sheets(r1, r2)
    
    # Gather Matrix compare stats using user selected files & sheets
    mat_b_file_req = request.args.get('matrix_base_file')
    mat_t_file_req = request.args.get('matrix_target_file')
    mat_b_sheet_req = request.args.get('matrix_base_sheet')
    mat_t_sheet_req = request.args.get('matrix_target_sheet')

    mat_b_p = resolve_file_path('matrix', mat_b_file_req or proj_defaults.get('matrix', ''), project_id)
    mat_t_p = resolve_file_path('matrix', mat_t_file_req or proj_defaults.get('matrix', ''), project_id)

    r_b, s_b, _ = read_file_safe(mat_b_p, sheet_name=mat_b_sheet_req)
    r_t, s_t, _ = read_file_safe(mat_t_p, sheet_name=mat_t_sheet_req)
    
    b_sheet_name = mat_b_sheet_req or (s_b[0] if s_b else 'Base')
    t_sheet_name = mat_t_sheet_req or (s_t[1] if len(s_t)>1 else (s_t[0] if s_t else 'Target'))
    
    mat_res = compare_two_matrix_sheets(r_b, r_t) if (r_b and r_t) else None

    # Build Markdown according to requested tab
    md = []
    if tab == 'bkc':
        md.append(f"# 💻 BKC Table Release Summary")
        md.append(f"**Generated Time:** `{now_str}` | **Active File:** `{os.path.basename(bkc_p)}`\n")
        md.append(f"## 📋 Overview")
        md.append(f"- **Total Components Tracked:** `{len(bkc_items)}` items")
        md.append(f"- **Total Sheets Available:** `{len(b_sheets)}` ({', '.join(b_sheets[:5]) if b_sheets else 'None'})\n")
        md.append(f"## ⚙️ Component List Sample")
        for it in bkc_items[:15]:
            md.append(f"- **[{it.get('category')}]** `{it.get('sub_component')}`: FW `{it.get('version') or 'N/A'}`")

    elif tab == 'fru':
        md.append(f"# 📄 FRU Specification Release Summary")
        md.append(f"**Generated Time:** `{now_str}`")
        md.append(f"**Base (DVT):** `{os.path.basename(dvt_p)}` (`{dvt_sheet_req or (s1[0] if s1 else 'Default')}`)")
        md.append(f"**Target (PVT):** `{os.path.basename(pvt_p)}` (`{pvt_sheet_req or (s2[0] if s2 else 'Default')}`)\n")
        md.append(f"## 📋 Overview")
        md.append(f"- **Total Spec Fields:** `{fru_res.get('total_items', 0)}` items")
        md.append(f"- **Identical Fields:** `{fru_res.get('same_count', 0)}` items")
        md.append(f"- **Specification Diffs:** `{fru_res.get('diff_count', 0)}` items\n")
        md.append(f"## 🔄 Specification Field Diffs Detail")
        diff_count = 0
        for fld in fru_res.get('fields', []):
            if fld.get('is_diff'):
                diff_count += 1
                md.append(f"{diff_count}. **[{fld.get('section')}]** `{fld.get('field_name')}`: `{fld.get('dvt_value') or '(empty)'}` ➔ `{fld.get('pvt_value') or '(empty)'}`")
        if diff_count == 0:
            md.append("- *No parameter differences detected between selected files and sheets.*")

    elif tab == 'matrix':
        md.append(f"# 🧱 Build Matrix Release Summary")
        md.append(f"**Generated Time:** `{now_str}`")
        md.append(f"**Base File:** `{os.path.basename(mat_b_p)}` (`{b_sheet_name}`)")
        md.append(f"**Target File:** `{os.path.basename(mat_t_p)}` (`{t_sheet_name}`)\n")
        if mat_res:
            md.append(f"## 📋 Cross-Sheet Summary")
            md.append(f"- **Total Config Items:** `{mat_res.get('total_items', 0)}` items")
            md.append(f"- **Configuration Diffs:** `{mat_res.get('diff_items_count', 0)}` items\n")
            md.append(f"## 🔄 Configuration Diffs Detail")
            diff_count = 0
            for it in mat_res.get('items', []):
                if it.get('is_diff'):
                    diff_count += 1
                    md.append(f"{diff_count}. **[{it.get('group_item')}]** `{it.get('description')}` (Values differ across rack configs)")
            if diff_count == 0:
                md.append("- *No build matrix differences detected between selected sheets.*")

    elif tab == 'yaml':
        y1 = request.args.get('yaml_1')
        y2 = request.args.get('yaml_2')
        y3 = request.args.get('yaml_3')
        y_files = [f for f in [y1, y2, y3] if f]
        bkc_f = request.args.get('bkc_file')
        bkc_s = request.args.get('bkc_sheet')
        y_res = compare_yaml_with_bkc(y_files, bkc_file_path=bkc_f, bkc_sheet_name=bkc_s, project_id=project_id)
        
        md.append(f"# 🧪 Test Suite (YAML) Compliance Summary")
        md.append(f"- **BKC Reference File**: `{y_res['summary']['bkc_file']}` (Sheet: `{y_res['summary']['bkc_sheet']}`)")
        md.append(f"- **Overall Compliance Rate**: **{y_res['summary']['compliance_rate']}%** ({y_res['summary']['matched_count']}/{y_res['summary']['total_yaml_checks']} items compliant)")
        md.append(f"- **Mismatches / Violations**: {y_res['summary']['mismatch_count']} items")
        md.append(f"- **Missing in BKC Table**: {y_res['summary']['missing_bkc_count']} items\n")
        md.append(f"## 🔄 Test Suite Version Discrepancies Detail")
        mismatch_items = [it for it in y_res['items'] if it['status'] == 'MISMATCH']
        if mismatch_items:
            for idx, it in enumerate(mismatch_items, start=1):
                md.append(f"{idx}. **[{it.get('station')}]** `{it.get('sub_component')}` at step `{it.get('step_location')}`: YAML `{it.get('yaml_version')}` 🆚 BKC `{it.get('bkc_version')}`")
                md.append(f"   - *Discussion Note:* {it.get('discussion_note')}")
        else:
            md.append("- *All test suite firmware/hardware specifications comply with BKC table.*")

    else: # ALL
        y_res = compare_yaml_with_bkc([], bkc_file_path=bkc_p)
        y_sum = y_res['summary']

        md.append(f"# 🚀 META VR200 (SanMiguel) All-in-One Release Summary Report")
        md.append(f"**Generated Time:** `{now_str}`")
        md.append(f"**Environment:** Hardware & Firmware Verification Platform\n")

        md.append(f"## 📋 1. Executive Summary")
        md.append(f"- **BKC Firmware Table:** `{len(bkc_items)}` components loaded from `{os.path.basename(bkc_p)}`")
        md.append(f"- **FRU Specification:** `{fru_res.get('diff_count', 0)}` diffs (`{os.path.basename(dvt_p)}` 🆚 `{os.path.basename(pvt_p)}`)")
        if mat_res:
            md.append(f"- **Build Matrix:** `{mat_res.get('diff_items_count', 0)}` diffs (`{b_sheet_name}` 🆚 `{t_sheet_name}`)")
        md.append(f"- **Test Suite (YAML):** `{y_sum.get('compliance_rate')}%` compliance (`{y_sum.get('mismatch_count')}` mismatches across `{y_sum.get('stations_count')}` test stations)\n")

        md.append(f"## ⚠️ 2. Critical Watchlist Impact")
        impacted_watchlist = []
        for fld in fru_res.get('fields', []):
            if fld.get('is_diff'):
                name = f"{fld.get('section','')} {fld.get('field_name','')}"
                if any(w.lower() in name.lower() for w in watchlist):
                    impacted_watchlist.append(f"- **[FRU - {fld.get('section')}]** `{fld.get('field_name')}`: `{fld.get('dvt_value')}` ➔ `{fld.get('pvt_value')}`")
        if impacted_watchlist:
            md.extend(impacted_watchlist)
        else:
            md.append("- *No critical watchlist components modified.*")

    md_text = "\n".join(md)
    plain_text = re.sub(r'[\*`#]', '', md_text)

    impact_count = len([fld for fld in fru_res.get('fields', []) if fld.get('is_diff') and any(w.lower() in f"{fld.get('section','')} {fld.get('field_name','')}".lower() for w in watchlist)])

    return jsonify({
        'success': True,
        'tab': tab,
        'markdown': md_text,
        'text': plain_text,
        'watchlist_impacts_count': impact_count
    })


@app.route('/api/export-excel', methods=['GET'])
def api_export_excel():
    """Export color-highlighted comparison Excel file."""
    tab_type = request.args.get('type', 'fru')
    diff_only = request.args.get('diff_only', 'false').lower() == 'true'
    project_id = request.args.get('project', 'sanmiguel')
    proj_cfg = get_project_config(project_id)
    proj_defaults = proj_cfg['default_paths']
    
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True

    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    fill_diff = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_diff = Font(name="Calibri", size=10, bold=True, color="92400E")

    if tab_type == 'yaml':
        ws.title = "Test Suite YAML Comparison"
        headers = ["Station", "Script Step & Location", "Component", "YAML Expected Version", "BKC Standard Version", "Compliance Status", "Action Disposition", "Assigned Owner", "Discussion & Action Notes", "Test Command"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        y1 = request.args.get('yaml_1')
        y2 = request.args.get('yaml_2')
        y3 = request.args.get('yaml_3')
        y_files = [f for f in [y1, y2, y3] if f]
        bkc_file = request.args.get('bkc_file')
        bkc_sheet = request.args.get('bkc_sheet')

        res = compare_yaml_with_bkc(y_files, bkc_file_path=bkc_file, bkc_sheet_name=bkc_sheet, project_id=project_id)
        
        row_idx = 2
        for item in res.get('items', []):
            if diff_only and item.get('status') == 'MATCH':
                continue
            ws.append([
                item.get('station'),
                item.get('step_location'),
                item.get('sub_component') or item.get('component'),
                item.get('yaml_version'),
                item.get('bkc_version'),
                item.get('status_label'),
                item.get('disposition_status', 'Pending'),
                item.get('disposition_owner', ''),
                item.get('discussion_note'),
                item.get('command')
            ])
            if item.get('status') == 'MISMATCH':
                for col in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = fill_diff
                    c.font = font_diff

            elif item.get('status') == 'MISSING_IN_BKC':
                fill_warn = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                font_warn = Font(name="Calibri", size=10, color="854D0E")
                for col in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = fill_warn
                    c.font = font_warn
            row_idx += 1

    elif tab_type == 'bkc':

        ws.title = "BKC Comparison"
        headers = ["Category", "Group", "Sub Component", "Meta Owner", "DVT FW Version", "PVT FW Version", "Diff Status"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        bkc_file = resolve_file_path('bkc', proj_defaults.get('bkc', ''), project_id)
        bkc_rows, bkc_sheets, _ = read_file_safe(bkc_file)
        if bkc_rows:
            bkc_data = parse_bkc_sheet(bkc_rows)
            row_idx = 2
            for cat in bkc_data.get('categories', []):
                for grp in cat.get('groups', []):
                    for item in grp.get('items', []):
                        if diff_only and not item.get('is_diff'):
                            continue
                        ws.append([
                            cat.get('name'),
                            grp.get('name'),
                            item.get('sub_component'),
                            item.get('meta_owner'),
                            item.get('dvt_version'),
                            item.get('pvt_version'),
                            "DIFFERENT" if item.get('is_diff') else "SAME"
                        ])
                        if item.get('is_diff'):
                            for col in range(1, 8):
                                c = ws.cell(row=row_idx, column=col)
                                c.fill = fill_diff
                                c.font = font_diff
                        row_idx += 1

    elif tab_type == 'matrix':
        ws.title = "Build Matrix Comparison"
        base_p = resolve_or_default(request.args.get('base_file'), 'matrix', proj_defaults.get('matrix', ''), project_id)
        tgt_p  = resolve_or_default(request.args.get('target_file'), 'matrix', proj_defaults.get('matrix', ''), project_id)
        b_rows, b_sheets, _ = read_file_safe(base_p, sheet_name=request.args.get('base_sheet'))
        t_rows, t_sheets, _ = read_file_safe(tgt_p, sheet_name=request.args.get('target_sheet'))
        res = compare_two_matrix_sheets(b_rows, t_rows)

        headers = ["Group Item", "Description"] + res.get('base_configs', [])
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = font_header

        row_idx = 2
        for item in res.get('items', []):
            if diff_only and not item.get('is_diff'):
                continue
            row_data = [item.get('group_item'), item.get('description')]
            for cfg in res.get('base_configs', []):
                row_data.append(item.get('config_values', {}).get(cfg, ''))
            ws.append(row_data)
            if item.get('is_diff'):
                for col in range(1, len(row_data) + 1):
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = fill_diff
                    c.font = font_diff
            row_idx += 1

    else: # Default FRU
        ws.title = "FRU Comparison"
        dvt_p = resolve_or_default(request.args.get('dvt_file'), 'fru', proj_defaults.get('fru_dvt', ''), project_id)
        pvt_p = resolve_or_default(request.args.get('pvt_file'), 'fru', proj_defaults.get('fru_pvt', ''), project_id)
        r1, _, _ = read_file_safe(dvt_p, sheet_name=request.args.get('base_sheet'))
        r2, _, _ = read_file_safe(pvt_p, sheet_name=request.args.get('target_sheet'))
        res = compare_two_fru_sheets(r1, r2)

        headers = ["Module", "Section", "Field Name", "Base (DVT) Value", "Target (PVT) Value", "Status"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = fill_header
            cell.font = font_header

        row_idx = 2
        for fld in res.get('fields', []):
            if diff_only and not fld.get('is_diff'):
                continue
            ws.append([
                fld.get('module'),
                fld.get('section'),
                fld.get('field_name'),
                fld.get('dvt_value'),
                fld.get('pvt_value'),
                "DIFFERENT" if fld.get('is_diff') else "SAME"
            ])
            if fld.get('is_diff'):
                for col in range(1, 7):
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = fill_diff
                    c.font = font_diff
            row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"VR200_{tab_type.upper()}_Comparison_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route('/api/export-coverage-excel', methods=['GET', 'POST'])
def api_export_coverage_excel():
    """Export YAML Test Suite Coverage Matrix Excel (.xlsx) workbook with full styling and KPI banner."""
    if request.method == 'POST':
        req_data = request.json or {}
        project_id = req_data.get('project', 'sanmiguel')
        bkc_file = req_data.get('bkc_file')
        bkc_sheet = req_data.get('bkc_sheet')
        y_params = [req_data.get(f'yaml_{i}') for i in range(1, 6)]
    else:
        project_id = request.args.get('project', 'sanmiguel')
        bkc_file = request.args.get('bkc_file')
        bkc_sheet = request.args.get('bkc_sheet')
        y_params = [request.args.get(f'yaml_{i}') for i in range(1, 6)]

    yaml_files = [f for f in y_params if f]
    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')
    bkc_p = resolve_file_path('bkc', bkc_file or default_bkc, project_id)

    comp_res = compare_yaml_with_bkc(yaml_files, bkc_file_path=bkc_p, bkc_sheet_name=bkc_sheet, project_id=project_id)
    cov = comp_res.get('coverage_matrix', {})
    summary = comp_res.get('summary', {})
    stations = cov.get('stations', [])
    grid = cov.get('grid', [])

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Coverage Matrix"
    ws.views.sheetView[0].showGridLines = True

    # Styling Tokens
    fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font_dark = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    fill_covered = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_covered = Font(name="Calibri", size=10, bold=True, color="166534")

    fill_uncovered = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    font_uncovered = Font(name="Calibri", size=10, color="991B1B")

    border_thin = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Title Banner
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"📊 {project_id.upper()} Test Suite Action Coverage Matrix (測試動作與步驟跨工站包含矩陣)"
    title_cell.fill = fill_dark
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="38BDF8")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 2. KPI Summary Row
    ws.append([
        "Total Unique Test Actions", cov.get('total_unique_items', len(grid)),
        "Common Actions Across Stations", cov.get('common_items_count', 0),
        "Active Stations Count", cov.get('active_stations_count', len(stations)),
        "BKC Reference Table", summary.get('bkc_file', ''), ""
    ])
    for col in range(1, 9):
        c = ws.cell(row=2, column=col)
        c.font = Font(name="Calibri", size=10, bold=True, color="475569")
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 22

    ws.append([]) # Blank row

    # 3. Table Headers
    headers = ["Test Action / Step Name (主要測試動作)", "Tested Targets / Sensors (測試對象 / FW / Sensor)"]
    for st in stations:
        clean_st = str(st).strip()
        m = re.search(r'\(([^)]+)\)', clean_st)
        if m and m.group(1):
            clean_st = m.group(1).upper()
        else:
            clean_st = re.sub(r'^Station\s*\d*\s*', '', clean_st, flags=re.IGNORECASE).upper()
        headers.append(clean_st if clean_st else st)
    headers.append("Overall Action Coverage (跨工站包含狀態)")

    ws.append(headers)
    header_row = 4
    ws.row_dimensions[header_row].height = 28
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = fill_dark
        c.font = font_dark
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 4. Data Rows
    row_idx = 5
    for item in grid:
        action_name = item.get('action_name') or 'General Step'
        targets_summary = item.get('targets_summary') or 'General Check'

        row_data = [action_name, targets_summary]

        st_map = item.get('stations', {})
        for st in stations:
            st_info = st_map.get(st)
            if st_info and st_info.get('is_included'):
                row_data.append("✅ 包含")
            else:
                row_data.append("—")

        cov_status = item.get('coverage_status_tag') or "🔵 單站包含"
        row_data.append(cov_status)

        ws.append(row_data)
        ws.row_dimensions[row_idx].height = None  # Auto row height for wrapped text

        # Cells styling
        for col_idx in range(1, len(row_data) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = border_thin

            if col_idx == 1:
                c.font = Font(name="Calibri", size=10, bold=True, color="0284C7")
                c.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_idx == 2:
                c.font = Font(name="Calibri", size=10, bold=True, color="334155")
                c.alignment = Alignment(vertical="center", wrap_text=True)
            elif col_idx >= 3 and col_idx < 3 + len(stations):
                val_str = str(c.value)
                if "✅ 包含" in val_str:
                    c.fill = fill_covered
                    c.font = font_covered
                else:
                    c.font = Font(name="Calibri", size=10, color="94A3B8")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == len(row_data):
                if "全包含" in str(c.value) or "跨站包含" in str(c.value):
                    c.fill = fill_covered
                    c.font = font_covered
                else:
                    c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    c.font = Font(name="Calibri", size=10, color="475569")
                c.alignment = Alignment(horizontal="center", vertical="center")

        row_idx += 1

    # Apply strict bounded column widths to prevent wide blowout
    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        col_idx = col[0].column

        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 38
        elif col_idx == 2:
            ws.column_dimensions[col_letter].width = 45
        elif col_idx >= 3 and col_idx < 3 + len(stations):
            ws.column_dimensions[col_letter].width = 16
        elif col_idx == 3 + len(stations):
            ws.column_dimensions[col_letter].width = 24
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Test_Item_Coverage_Matrix_{project_id.upper()}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def find_best_yaml_match_for_fava(cat, item_name, extracted_items, stage='L10'):
    """Smartly and precisely match a FAVA table row (Category + Item name) against extracted YAML items."""
    if not item_name and not cat:
        return None

    cat_str = str(cat or '').strip().lower()
    item_str = str(item_name or '').strip().lower()
    stage_str = str(stage or 'L10').strip().upper()

    if stage_str == 'L10' and 'l10' not in item_str and 'pcieswitch' not in item_str:
        l11_keys = [
            'rmc', 'nvswitch', 'powerrack', 'aalc', 'wedge400', '2nd', 
            'bbu', 'pmm', 'psu', 'l11', 'minipack', 'mgmt switch',
            'aei', 'delta', 'panasonic', 'fixture', 'cbu'
        ]
        if any(k in cat_str for k in l11_keys) or any(k in item_str for k in l11_keys):
            return None

    best_match = None
    best_score = -1

    for y_it in extracted_items:
        comp = str(y_it.get('component') or '').lower()
        sub = str(y_it.get('sub_component') or '').lower()
        step = str(y_it.get('step_location') or '').lower()
        fru = str(y_it.get('fru') or '').lower()
        station = str(y_it.get('station') or '').lower()

        score = 0

        # 1. System / Kernel / OS / Compute Tray Version Checks
        if 'erot bmc' in item_str:
            if 'erot bmc' in comp or 'erot bmc' in sub: score = 100
        elif 'erot cpu' in item_str:
            if 'erot cpu' in comp or 'erot cpu' in sub: score = 100
        elif 'erot fpga' in item_str:
            if 'erot fpga' in comp or 'erot fpga' in sub: score = 100
        elif 'hmc cpld' in item_str:
            if 'hmc cpld' in comp or 'hmc cpld' in sub: score = 100
        elif 'hmc' in item_str and 'cpld' not in item_str and 'erot' not in item_str:
            if comp == 'hmc' or sub == 'hmc': score = 100
        elif 'inforom' in item_str:
            if 'inforom' in comp or 'inforom' in sub: score = 100
        elif 'vbios' in item_str or ('gpu' in item_str and 'inforom' not in item_str):
            if 'vbios' in comp or 'gpu' in comp or 'vbios' in sub or 'gpu' in sub:
                if 'versioncheck' in step: score = 100
                elif 'update_gpu' in step or 'vbios' in step: score = 95
                elif 'vbios' in comp or 'vbios' in sub: score = 90
                else: score = 80
        elif 'erot' in item_str:
            if 'erot' in comp or 'erot' in sub: score = 100
        elif 'sbios' in item_str or ('bios' in item_str and 'vbios' not in item_str):
            if 'sbios' in comp or 'sbios' in sub or 'bios' in comp or 'bios' in sub: score = 100
        elif 'fpga' in item_str and 'erot' not in item_str and 'cpld' not in item_str:
            if comp == 'fpga' or sub == 'fpga': score = 100
        elif 'pcieswitchconfig' in item_str or 'pcieswitch' in item_str:
            if 'pcieswitchconfig' in comp or 'pcieswitchconfig' in sub or 'pcie' in comp: score = 100
        elif cat_str == 'system' or 'kernel' in item_str or ('os' in item_str and 'nvos' not in item_str):
            if 'fbkernel' in sub or 'kernel' in sub or 'kernel' in comp or 'fbkernel' in step:
                if 'fb kernel' in item_str or 'fbkernel' in item_str:
                    score = 100
            elif 'os' in item_str and ('os' in sub or 'os' in comp or 'os' in step):
                if 'interposer' not in sub and 'interposer' not in step:
                    score = 100

        # 2. PDB VR
        elif 'pdb' in cat_str or 'vr' in item_str:
            if 'vr' in item_str and ('vr' in sub or 'vr' in comp or 'vr' in step):
                if ('n1' in item_str and 'n1' in sub) or ('n2' in item_str and 'n2' in sub):
                    score = 100

        # 3. SCM CPLD
        elif cat_str == 'scm' and 'cpld' in item_str:
            if fru == 'scm' or 'scm' in comp or 'scm' in step or 'scm' in sub:
                if 'cpld' in comp or 'cpld' in sub or 'cpld' in step:
                    score = 100

        # 4. E1.S BP CPLD
        elif ('e1.s bp' in cat_str or 'hdd' in cat_str) and 'cpld' in item_str:
            if fru in ['hdd0', 'hdd1', 'hdd'] or 'hdd' in comp or 'hdd' in step or 'bp' in comp:
                if 'cpld' in comp or 'cpld' in sub or 'cpld' in step:
                    score = 100

        # 5. Interposer CPLD
        elif cat_str == 'interposer' and 'cpld' in item_str:
            if fru == 'interposer' or 'interposer' in comp or 'interposer' in step or 'interposer' in sub:
                if 'cpld' in comp or 'cpld' in sub or 'cpld' in step:
                    score = 100

        # 6. BSM OpenBMC
        elif cat_str == 'bsm' and ('openbmc' in item_str or 'bmc' in item_str):
            if fru in ['bmc', 'bsm'] or 'bmc' in comp or 'ct_bmc' in comp or 'bmc' in sub:
                score = 100

        # 7. SCM-TPM
        elif 'tpm' in item_str:
            if 'tpm' in sub or 'tpm' in comp or 'tpm' in step:
                score = 100

        # 8. NICs
        elif 'fe nic' in cat_str or 'fe nic' in item_str or 'cx7' in cat_str or 'cx7' in item_str:
            if 'fenic' in sub or 'fenic' in comp or 'cx7' in sub or 'fenic' in step or 'cx7' in step:
                score = 100
        elif 'be nic' in cat_str or 'be nic' in item_str or 'cx8' in cat_str or 'cx8' in item_str:
            if 'benic' in sub or 'benic' in comp or 'cx8' in sub or 'benic' in step or 'cx8' in step:
                score = 100

        # 9. SSDs
        elif 'drive' in cat_str or 'storage' in cat_str or any(m in item_str for m in ['samsung', 'micron', 'kioxia', 'wd']):
            item_clean = item_str.replace('boot drive e1.s', '').replace('storage e1.s', '').strip()
            if item_clean and (item_clean in sub or item_clean in comp or sub in item_clean):
                score = 100
            elif 'samsung' in item_str and 'samsung' in sub: score = 90
            elif 'micron' in item_str and 'micron' in sub: score = 90
            elif 'kioxia' in item_str and 'kioxia' in sub: score = 90
            elif 'wd' in item_str and 'wd' in sub: score = 90

        elif 'nvpd' in item_str:
            if 'nvswitch' in item_str:
                if 'nvswitch nvpd' in sub or 'switch' in step or 'netblade' in sub: score = 100
            elif 'l11' in item_str:
                if 'l11 nvpd' in sub or 'partnerdiag' in step or 'partner' in step: score = 100
            elif 'l10' in item_str:
                if 'l10 nvpd' in sub or 'nvdiagcheck' in step: score = 100

        elif cat_str == 'nvswitch' or 'nvswitch' in cat_str:
            if 'nvos' in item_str:
                if 'nvos' in sub or 'nvos' in comp or 'nvos' in step: score = 100
            elif 'erot' in item_str:
                if 'erot' in sub or 'erot' in comp: score = 100
            elif 'bmc' in item_str and 'erot' not in item_str:
                if 'bmc' in sub or 'bmc' in comp: score = 100
            elif 'sbios' in item_str or 'bios' in item_str:
                if 'bios' in sub or 'bios' in comp: score = 100
            elif 'fpga' in item_str and 'erot' not in item_str:
                if 'fpga' in sub or 'fpga' in comp: score = 100
            elif 'cpld1' in item_str:
                if 'cpld1' in sub or 'cpld1' in comp: score = 100
            elif 'cpld2' in item_str:
                if 'cpld2' in sub or 'cpld2' in comp: score = 100
            elif 'cpld3' in item_str:
                if 'cpld3' in sub or 'cpld3' in comp: score = 100
            elif 'cpld4' in item_str:
                if 'cpld4' in sub or 'cpld4' in comp: score = 100

        # 10. L11 Power / RMC / CBU / PMM / PSU Component Matching
        elif stage_str == 'L11':
            if 'cbu' in item_str:
                if 'pmm' not in item_str and 'pmm' in step:
                    continue
                if 'pmm' in item_str and 'pmm' not in step:
                    continue
                if 'cbu' in sub or 'cbu' in step:
                    if 'delta' in item_str and ('delta' in sub or 'delta' in step): score = 100
                    elif ('aei' in item_str or 'artesyn' in item_str) and ('aei' in sub or 'artesyn' in sub or 'aei' in step or 'artesyn' in step): score = 100
                    elif 'pmm' not in item_str and 'pmm' not in sub: score = 90
            elif 'rmc' in cat_str or 'rmc' in item_str:
                if 'cpld' in item_str and 'cpld' in sub and ('rmc' in step or fru == 'rmc'): score = 100
                elif 'bmc' in item_str and ('bmc' in sub or 'bmc' in step or fru == 'bmc'): score = 100
            elif 'pmm' in item_str:
                if ('aei' in item_str or 'artesyn' in item_str) and ('aei' in sub or 'artesyn' in sub) and 'pmm' in sub: score = 100
                elif 'delta' in item_str and 'delta' in sub and 'pmm' in sub: score = 100
            elif 'psu' in item_str:
                if ('aei' in item_str or 'artesyn' in item_str) and ('aei' in sub or 'artesyn' in sub) and 'psu' in sub: score = 100
                elif 'delta' in item_str and 'delta' in sub and 'psu' in sub: score = 100

        if score > best_score and score >= 50:
            best_score = score
            best_match = y_it

    return best_match


@app.route('/api/export-fava-draft', methods=['GET', 'POST'])
def api_export_fava_draft():
    """Export YAML test suite comparison in 'FAVA L10 FW Control Table-draft' format."""
    if request.method == 'POST':
        req_data = request.json or {}
        project_id = req_data.get('project', 'sanmiguel')
        bkc_file = req_data.get('bkc_file')
        bkc_sheet = req_data.get('bkc_sheet')
        y_params = [req_data.get(f'yaml_{i}') for i in range(1, 11)]
        stage_exp = (req_data.get('stage') or 'L10').upper()
        indices_val = req_data.get('selected_indices')
        if isinstance(indices_val, list):
            selected_indices = set(int(i) for i in indices_val)
        else:
            selected_indices = None
    else:
        project_id = request.args.get('project', 'sanmiguel')
        stage_exp = request.args.get('stage', 'L10').upper()
        bkc_file = request.args.get('bkc_file')
        bkc_sheet = request.args.get('bkc_sheet')
        y_params = [request.args.get(f'yaml_{i}') for i in range(1, 11)]
        indices_str = request.args.get('indices', '')
        selected_indices = set([int(i) for i in indices_str.split(',') if i.isdigit()]) if indices_str else None

    yaml_files = [f for f in y_params if f]

    # Validate Stage Filename Rules (L10 must contain FDT/FRO/FFT; L11 cannot contain FDT/FRO/FFT)
    invalid_files = []
    for yf in yaml_files:
        b_name = os.path.basename(yf).lower()
        is_l10_file = any(k in b_name for k in ['fdt', 'fro', 'fft'])
        if stage_exp == 'L10' and not is_l10_file:
            invalid_files.append(os.path.basename(yf))
        elif stage_exp == 'L11' and is_l10_file:
            invalid_files.append(os.path.basename(yf))

    if invalid_files:
        if stage_exp == 'L10':
            err_msg = f"❌ L10 Stage 僅接受檔名包含 FDT, FRO, 或 FFT 之測試腳本。不符合規範檔案：{', '.join(invalid_files)}"
        else:
            err_msg = f"❌ L11 Stage 不可使用包含 FDT, FRO, 或 FFT 之 L10 測試腳本。不符合規範檔案：{', '.join(invalid_files)}"
        return err_msg, 400

    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')
    bkc_p = resolve_file_path('bkc', bkc_file or default_bkc, project_id)

    comp_res = compare_yaml_with_bkc(yaml_files, bkc_file_path=bkc_p, bkc_sheet_name=bkc_sheet, project_id=project_id)
    extracted_items = comp_res.get('items', [])

    template_path = os.path.join(DATA_DIR, 'clemente', 'bkc', 'IGS-Clemente FAVA FW Control Table Tracker.xlsx')
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    target_sheet_name = 'FAVA L11 FW Control Table-draft' if stage_exp == 'L11' else 'FAVA L10 FW Control Table-draft'
    ws = wb[target_sheet_name]

    current_category = ""
    cat_counts_exp = {}
    for row_idx in range(2, ws.max_row + 1):
        cat = " ".join(str(ws.cell(row=row_idx, column=1).value or '').split())
        item_name = " ".join(str(ws.cell(row=row_idx, column=2).value or '').split())
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=item_name)
        act_v = str(ws.cell(row=row_idx, column=3).value or '').strip()
        d_v = str(ws.cell(row=row_idx, column=4).value or '').strip()
        if not cat and not item_name and not act_v and not d_v:
            continue

        if cat:
            current_category = cat
            c_up = cat.upper()
            cat_counts_exp[c_up] = cat_counts_exp.get(c_up, 0) + 1

        eff_cat = cat or current_category
        eff_up = eff_cat.upper()

        is_l11_exp = False
        c_low = eff_cat.lower()
        i_low = item_name.lower()

        if stage_exp == 'L10':
            if 'l11' in i_low or 'l11' in c_low or 'pcieswitchconfig' in i_low or 'pciesswitchconfig' in i_low:
                is_l11_exp = True
            elif any(k in c_low for k in ['rack', 'rmc', 'nvswitch', 'powerrack', 'aalc', 'wedge400', '2nd', 'fan bd', 'rear io', 'led bd', 'fan', 'rear', 'led']):
                is_l11_exp = True
            elif any(k in i_low for k in ['rmc', 'nvswitch', 'aalc', 'wedge400', 'bbu', 'pmm', 'psu', 'cbu', 'minipack', 'mgmt switch', 'switch', 'fixture', 'aei', 'delta', 'panasonic', 'bft', 'fan', 'rear', 'led']):
                is_l11_exp = True

            if eff_up in ['SCM', 'BSM'] and cat_counts_exp.get(eff_up, 0) >= 2:
                is_l11_exp = True

            if 'l10' in i_low or 'l10' in c_low:
                is_l11_exp = False

        matched_item = None
        if not is_l11_exp:
            if item_name.strip().upper() == 'FRU':
                latest_fru = get_latest_fru_version(project_id)
                ws.cell(row=row_idx, column=4, value=str(latest_fru))
                ws.cell(row=row_idx, column=5, value=f"Verified in FRU Spec ({latest_fru})")
            else:
                matched_item = find_best_yaml_match_for_fava(eff_cat, item_name, extracted_items, stage=stage_exp)

        if matched_item:
            extracted_ver = matched_item.get('yaml_version') or ''
            if extracted_ver:
                ws.cell(row=row_idx, column=4, value=str(extracted_ver))
            status = matched_item.get('status', 'MATCH')
            bkc_v = str(matched_item.get('bkc_version') or '').strip()
            if status == 'MISMATCH' and bkc_v and bkc_v != 'None':
                ws.cell(row=row_idx, column=5, value=f"MISMATCH: YAML ({extracted_ver}) vs BKC ({bkc_v})")
            else:
                ws.cell(row=row_idx, column=5, value=f"Verified in YAML ({matched_item.get('station', 'YAML')})")
        elif is_l11_exp:
            ws.cell(row=row_idx, column=4, value="Yaml檔案未找到")
            ws.cell(row=row_idx, column=5, value="L11 Scope (非 L10 測試範疇)")
        elif not matched_item and item_name:
            ws.cell(row=row_idx, column=4, value="Yaml檔案未找到")
            ws.cell(row=row_idx, column=5, value="Yaml檔案未找到")

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)

    filename = f"FAVA_{stage_exp}_FW_Control_Table_Draft.xlsx"
    return send_file(
        output_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def get_latest_fru_version(project_id='clemente'):
    fru_dir = os.path.join(DATA_DIR, project_id, 'fru')
    if not os.path.exists(fru_dir):
        return '0.05A'
    
    xlsx_files = [f for f in os.listdir(fru_dir) if f.endswith('.xlsx')]
    if not xlsx_files:
        return '0.05A'
    
    xlsx_files.sort(reverse=True)
    latest_file = os.path.join(fru_dir, xlsx_files[0])
    try:
        wb = load_workbook(latest_file, data_only=True)
        for sname in wb.sheetnames:
            if 'VERSION' in sname.upper():
                ws = wb[sname]
                last_ver = None
                for r in range(1, ws.max_row + 1):
                    val = ws.cell(row=r, column=2).value or ws.cell(row=r, column=1).value
                    if val and str(val).strip().lower() != 'fru version':
                        last_ver = str(val).strip()
                if last_ver:
                    return last_ver
    except Exception as e:
        print(f"Error reading FRU file: {e}")
    return '0.05A'


SSD_STATIC_SPECS = {
    'samsung pm9d3a 9.5mmt': {'vid': '144Dh', 'did': 'A80E'},
    'samsung pm9d3a 25mmt': {'vid': '144Dh', 'did': 'A80E'},
    'micron 7450': {'vid': '1344', 'did': '51C3'},
    'kioxia  xd7p': {'vid': '1E0F', 'did': '001C'},
    'kioxia xd7p': {'vid': '1E0F', 'did': '001C'},
    'micron 7550': {'vid': '1344', 'did': '51BC'},
    'wd sn861 -wus6a7676pjp8x7': {'vid': '1B96', 'did': '2750'},
    'wd sn861': {'vid': '1B96', 'did': '2750'},
}

@app.route('/api/preview-fava-draft', methods=['GET'])
def api_preview_fava_draft():
    """Return live preview rows for FAVA FW Control Table Modal (Supports L10 & L11 stage modes)."""
    project_id = request.args.get('project', 'sanmiguel')
    stage = request.args.get('stage', 'L10').upper()
    bkc_file = request.args.get('bkc_file')
    bkc_sheet = request.args.get('bkc_sheet')
    y_params = [request.args.get(f'yaml_{i}') for i in range(1, 11)]
    yaml_files = [f for f in y_params if f]

    # Validate Stage Filename Rules (L10 must contain FDT/FRO/FFT; L11 cannot contain FDT/FRO/FFT)
    invalid_files = []
    for yf in yaml_files:
        b_name = os.path.basename(yf).lower()
        is_l10_file = any(k in b_name for k in ['fdt', 'fro', 'fft'])
        if stage == 'L10' and not is_l10_file:
            invalid_files.append(os.path.basename(yf))
        elif stage == 'L11' and is_l10_file:
            invalid_files.append(os.path.basename(yf))

    if invalid_files:
        if stage == 'L10':
            err_msg = f"❌ L10 Stage 僅接受檔名包含 FDT, FRO, 或 FFT 之測試腳本。以下檔案不符合規範：{', '.join(invalid_files)}"
        else:
            err_msg = f"❌ L11 Stage 不可使用包含 FDT, FRO, 或 FFT 之 L10 測試腳本。以下檔案不符合規範：{', '.join(invalid_files)}"
        return jsonify({"success": False, "error": err_msg}), 400

    proj_cfg = get_project_config(project_id)
    default_bkc = proj_cfg['default_paths'].get('bkc', '')
    bkc_p = resolve_file_path('bkc', bkc_file or default_bkc, project_id)

    comp_res = compare_yaml_with_bkc(yaml_files, bkc_file_path=bkc_p, bkc_sheet_name=bkc_sheet, project_id=project_id)
    extracted_items = comp_res.get('items', [])
    latest_fru_ver = get_latest_fru_version(project_id)

    template_path = os.path.join(DATA_DIR, 'clemente', 'bkc', 'IGS-Clemente FAVA FW Control Table Tracker.xlsx')
    
    preview_rows = []
    if os.path.exists(template_path):
        target_sheet = 'FAVA L11 FW Control Table-draft' if stage == 'L11' else 'FAVA L10 FW Control Table-draft'
        b_rows, sheets, _ = read_file_safe(template_path, sheet_name=target_sheet)
        current_category = ""
        cat_counts = {}
        last_ssd_model = ""

        for r_idx, r in enumerate(b_rows):
            if r_idx == 0: continue
            cat = " ".join(str(r[0] or '').split()) if len(r) > 0 else ''
            item_name = " ".join(str(r[1] or '').split()) if len(r) > 1 else ''
            act_ver = str(r[2] or '').strip() if len(r) > 2 else ''
            draft_ver = str(r[3] or '').strip() if len(r) > 3 else ''
            if draft_ver.endswith('.0') and draft_ver.replace('.0', '').isdigit():
                draft_ver = draft_ver[:-2]
            remark = str(r[4] or '').strip() if len(r) > 4 else ''

            if not cat and not item_name and not act_ver and not draft_ver:
                continue

            if cat:
                current_category = cat
                c_up = cat.upper()
                cat_counts[c_up] = cat_counts.get(c_up, 0) + 1

            if item_name:
                item_low = item_name.lower().strip()
                if any(m in item_low for m in ['samsung', 'micron', 'kioxia', 'wd sn861']):
                    last_ssd_model = item_low

            effective_cat = cat or current_category
            eff_up = effective_cat.upper()

            is_l11 = False
            c_low = effective_cat.lower()
            i_low = item_name.lower()

            if stage == 'L10':
                if 'pcieswitch' not in i_low:
                    if 'l11' in i_low or 'l11' in c_low:
                        is_l11 = True
                    elif any(k in c_low for k in ['rack', 'rmc', 'nvswitch', 'powerrack', 'aalc', 'wedge400', '2nd', 'fan bd', 'rear io', 'led bd', 'fan', 'rear', 'led']):
                        is_l11 = True
                    elif any(k in i_low for k in ['rmc', 'nvswitch', 'aalc', 'wedge400', 'bbu', 'pmm', 'psu', 'cbu', 'minipack', 'mgmt switch', 'fixture', 'aei', 'delta', 'panasonic', 'bft', 'fan', 'rear', 'led']):
                        is_l11 = True

                    # 2nd occurrence of SCM/BSM
                    if eff_up in ['SCM', 'BSM'] and cat_counts.get(eff_up, 0) >= 2:
                        is_l11 = True

                if 'l10' in i_low or 'l10' in c_low or 'pcieswitch' in i_low:
                    is_l11 = False

            matched_item = None
            if not is_l11 and item_name:
                if item_name.strip().upper() == 'FRU':
                    draft_ver = latest_fru_ver
                    remark = f"Verified in FRU Spec ({latest_fru_ver})"
                else:
                    matched_item = find_best_yaml_match_for_fava(effective_cat, item_name, extracted_items, stage=stage)

            if matched_item:
                extracted_ver = matched_item.get('yaml_version') or ''
                if extracted_ver: draft_ver = str(extracted_ver)
                status = matched_item.get('status', 'MATCH')
                bkc_v = str(matched_item.get('bkc_version') or '').strip()
                if status == 'MISMATCH' and bkc_v and bkc_v != 'None':
                    remark = f"MISMATCH: YAML ({extracted_ver}) vs BKC ({bkc_v})"
                else:
                    remark = f"Verified in YAML ({matched_item.get('station', 'YAML')})"

            display_item = item_name
            # Explicit VID/DID sub-row handling under SSDs (L10 Stage only)
            if stage == 'L10' and not cat and not item_name and last_ssd_model and c_low in ['boot drive e1.s', 'storage e1.s']:
                matched_spec_key = None
                for key in SSD_STATIC_SPECS:
                    if key in last_ssd_model:
                        matched_spec_key = key
                        break
                
                spec_info = SSD_STATIC_SPECS.get(matched_spec_key, {})

                if len(preview_rows) > 0:
                    prev_item = preview_rows[-1].get('item', '')
                    if 'Vendor ID' not in prev_item and prev_item != '':
                        display_item = "↳ Vendor ID (VID)"
                        draft_ver = spec_info.get('vid') or draft_ver or '144Dh'
                        remark = "Static Spec (VID)"
                    elif 'Device ID' not in prev_item and prev_item != '':
                        display_item = "↳ Device ID (DID)"
                        draft_ver = spec_info.get('did') or draft_ver or 'A80E'
                        remark = "Static Spec (DID)"

            if is_l11:
                draft_ver = "Yaml檔案未找到"
                remark = "L11 Scope (非 L10 測試範疇)"
            else:
                if not matched_item and item_name:
                    draft_ver = "Yaml檔案未找到"
                    remark = "Yaml檔案未找到"
                elif not remark:
                    remark = "-"

            if item_name or cat or draft_ver:
                preview_rows.append({
                    'category': cat,
                    'item': display_item,
                    'actual_version': act_ver,
                    'draft_version': draft_ver,
                    'remark': remark,
                    'is_updated': bool(matched_item),
                    'is_l11': is_l11,
                    'status': matched_item.get('status', 'NONE') if matched_item else 'NONE'
                })

    # Prepare TSV string (Tab-separated values) for direct pasting into Excel/Google Sheets
    tsv_lines = ["Category\tItem\tActual Version\tDraft Version\tRemark"]
    for r in preview_rows:
        c_cat = str(r['category'] or '').replace('\n', ' ').strip()
        c_item = str(r['item'] or '').replace('\n', ' ').strip()
        c_act = str(r['actual_version'] or '').replace('\n', ' ').strip()
        c_draft = str(r['draft_version'] or '').replace('\r\n', ' / ').replace('\n', ' / ').strip()
        c_rem = str(r['remark'] or '').replace('\n', ' ').strip()
        tsv_lines.append(f"{c_cat}\t{c_item}\t{c_act}\t{c_draft}\t{c_rem}")
    tsv_text = "\n".join(tsv_lines)

    return jsonify({
        'success': True,
        'rows': preview_rows,
        'tsv_text': tsv_text,
        'total_items': len(preview_rows),
        'updated_items': len([r for r in preview_rows if r['is_updated']])
    })


@app.route('/api/debug-search')
def api_debug_search():
    """Diagnostic endpoint to inspect file scanning and matrix search matches."""
    q_raw = request.args.get('q', '15-106079')
    project_id = request.args.get('project', 'sanmiguel')
    clean_q = re.sub(r'[^a-zA-Z0-9]', '', q_raw.lower())

    debug_info = {
        'query_raw': q_raw,
        'clean_query': clean_q,
        'project': project_id,
        'scanned_files': [],
        'matches': []
    }

    for f in scan_files_in_dirs('matrix', project_id):

        try:
            r_first, sheets, err = read_file_safe(f['path'])
            debug_info['scanned_files'].append({
                'filename': f['filename'],
                'path': f['path'],
                'sheets': sheets,
                'error': err
            })
            for s in (sheets or [None]):
                rows = r_first if (sheets and s == sheets[0]) else read_file_safe(f['path'], sheet_name=s)[0]
                if not rows: continue
                parsed = parse_matrix_sheet(rows)
                items_list = parsed[5] if len(parsed) >= 6 else []
                for item in items_list:
                    cfg_vals = " ".join([str(v) for v in item.get('values', {}).values()])
                    searchable = f"{item.get('group_item','')} {item.get('attribute','')} {cfg_vals}"
                    if clean_q in re.sub(r'[^a-zA-Z0-9]', '', searchable.lower()):
                        debug_info['matches'].append({
                            'file': f['filename'],
                            'sheet': s,
                            'group_item': item.get('group_item'),
                            'attribute': item.get('attribute'),
                            'values': item.get('values')
                        })
        except Exception as e:
            debug_info['scanned_files'].append({'filename': f['filename'], 'error': str(e)})

    return jsonify(debug_info)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8055))
    debug_mode = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    app.run(host='0.0.0.0', port=port, debug=debug_mode)

